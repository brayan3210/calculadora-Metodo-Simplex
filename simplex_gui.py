"""
============================================================================
 SIMPLEX GUI - Resolución de Programación Lineal por Método Simplex (Gran M)
 Con interfaz gráfica, parser de lenguaje matemático y exportación a Excel
============================================================================

USO:
    1) Ejecutar:    python3 simplex_gui.py
    2) Pegar el modelo matemático en la caja de texto (formato libre).
    3) Pulsar "Resolver" - se muestran las tablas paso a paso.
    4) Pulsar "Exportar a Excel" - genera un .xlsx con todo el procedimiento.

REQUISITOS:
    - Python 3.8+
    - tkinter (incluido en Python estándar)
    - openpyxl (instalar con:  pip install openpyxl)

FORMATO ACEPTADO DEL MODELO (es flexible):
    Max  Z = 22000 X1 + 20000 X2
    s.a.
       2,5 X1 + 3 X2  <=  4500
       3 X1 + 6 X2    <=  8400
       14 X1 + 10 X2  <=  20000
       X1 + X2        <=  1700
       X1             >=  600
       X1, X2 >= 0

PARA HACER EJECUTABLE .EXE (opcional):
    pip install pyinstaller
    pyinstaller --onefile --windowed simplex_gui.py
    (el .exe queda en la carpeta  dist/  )
============================================================================
"""

import re
import os
import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox
from fractions import Fraction
from copy import deepcopy
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# La llamada a Groq se hace con urllib (libreria estándar de Python).
# No requiere instalar nada extra.
import urllib.request
import urllib.error
import json as _json

CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".simplex_gui_config.json")


def cargar_config():
    try:
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            return _json.load(f)
    except Exception:
        return {}


def guardar_config(cfg):
    try:
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            _json.dump(cfg, f)
    except Exception:
        pass


# ============================================================================
#  Coeficiente simbólico a + b*M  (M -> +infinito)
# ============================================================================
class Coef:
    __slots__ = ("const", "m")

    def __init__(self, const=0, m_coef=0):
        self.const = const if isinstance(const, Fraction) else Fraction(const)
        self.m = m_coef if isinstance(m_coef, Fraction) else Fraction(m_coef)

    def _c(self, o):  return o if isinstance(o, Coef) else Coef(o)

    def __add__(self, o):
        o = self._c(o); return Coef(self.const + o.const, self.m + o.m)
    def __radd__(self, o): return self.__add__(o)
    def __sub__(self, o):
        o = self._c(o); return Coef(self.const - o.const, self.m - o.m)
    def __rsub__(self, o): return self._c(o).__sub__(self)

    def __mul__(self, o):
        if isinstance(o, Coef):
            if self.m != 0 and o.m != 0:
                raise ValueError("M^2 no soportado")
            return Coef(self.const * o.const,
                        self.const * o.m + self.m * o.const)
        f = o if isinstance(o, Fraction) else Fraction(o)
        return Coef(self.const * f, self.m * f)
    def __rmul__(self, o): return self.__mul__(o)

    def __truediv__(self, o):
        if isinstance(o, Coef):
            if o.m != 0: raise ValueError("División por M no soportada")
            d = o.const
        else:
            d = o if isinstance(o, Fraction) else Fraction(o)
        return Coef(self.const / d, self.m / d)

    def __neg__(self): return Coef(-self.const, -self.m)
    def __eq__(self, o):
        o = self._c(o); return self.const == o.const and self.m == o.m
    def __hash__(self): return hash((self.const, self.m))
    def es_cero(self): return self.const == 0 and self.m == 0

    def mas_positivo_que(self, o):
        if self.m != o.m: return self.m > o.m
        return self.const > o.const
    def mas_negativo_que(self, o):
        if self.m != o.m: return self.m < o.m
        return self.const < o.const

    @staticmethod
    def _fmt(x):
        if x.denominator == 1: return str(x.numerator)
        return f"{x.numerator}/{x.denominator}"

    def __str__(self):
        if self.m == 0: return self._fmt(self.const)
        if self.m == 1:    m_only = "M"
        elif self.m == -1: m_only = "-M"
        else:              m_only = f"{self._fmt(self.m)}M"
        if self.const == 0: return m_only
        c_part = self._fmt(self.const)
        if self.m > 0:
            sep = "+"; m_str = "M" if self.m == 1 else f"{self._fmt(self.m)}M"
        else:
            sep = "-"; m_str = "M" if self.m == -1 else f"{self._fmt(-self.m)}M"
        return f"{c_part}{sep}{m_str}"

    # Valor numérico aproximado con M grande (para exportar a Excel)
    def num_aprox(self, M=1e9):
        return float(self.const) + float(self.m) * M


def fmt_num(x):
    if isinstance(x, Coef):  return str(x)
    if isinstance(x, Fraction):
        if x.denominator == 1: return str(x.numerator)
        return f"{x.numerator}/{x.denominator}"
    if isinstance(x, float):
        return fmt_num(Fraction(x).limit_denominator(10000))
    return str(x)


# ============================================================================
#  SOLVER SIMPLEX  (registra todas las iteraciones para luego mostrarlas)
# ============================================================================
class SimplexSolver:

    def __init__(self, tipo, c, restricciones, nombres_var=None, max_iter=50):
        if tipo.lower() not in ("max", "min"):
            raise ValueError("tipo debe ser 'max' o 'min'")
        self.tipo = tipo.lower()
        self.c_orig = [Fraction(v) for v in c]
        self.n_dec = len(c)

        rests = []
        for a, s, b in restricciones:
            if len(a) != self.n_dec:
                raise ValueError("Coeficientes incompatibles en una restricción.")
            if s not in ("<=", ">=", "="):
                raise ValueError(f"Signo no válido: {s}")
            b_f = Fraction(b)
            a_f = [Fraction(v) for v in a]
            if b_f < 0:
                a_f = [-v for v in a_f]; b_f = -b_f
                if s == "<=":   s = ">="
                elif s == ">=": s = "<="
            rests.append((a_f, s, b_f))
        self.restricciones = rests
        self.nombres_dec = nombres_var or [f"X{i+1}" for i in range(self.n_dec)]
        self.max_iter = max_iter
        self._construir_estandar()

    def _construir_estandar(self):
        info_rest = []
        idx = self.n_dec
        for i, (_, s, _) in enumerate(self.restricciones):
            sub = i + 1
            extras = []
            if s == "<=":
                extras.append(("S+", idx, sub)); idx += 1
            elif s == ">=":
                extras.append(("S-", idx, sub)); idx += 1
                extras.append(("A",  idx, sub)); idx += 1
            elif s == "=":
                extras.append(("A",  idx, sub)); idx += 1
            info_rest.append(extras)
        self.n_total = idx

        self.nombres_var = list(self.nombres_dec)
        self.tipo_var = ["dec"] * self.n_dec
        for extras in info_rest:
            for tp, _, sub in extras:
                if tp == "S+":
                    self.nombres_var.append(f"S{sub}"); self.tipo_var.append("S+")
                elif tp == "S-":
                    self.nombres_var.append(f"S{sub}"); self.tipo_var.append("S-")
                elif tp == "A":
                    self.nombres_var.append(f"A{sub}"); self.tipo_var.append("A")

        self.C = []
        for j in range(self.n_total):
            tp = self.tipo_var[j]
            if tp == "dec":            self.C.append(Coef(self.c_orig[j]))
            elif tp in ("S+", "S-"):   self.C.append(Coef(0))
            elif tp == "A":
                self.C.append(Coef(0, -1) if self.tipo == "max" else Coef(0, 1))

        self.A = []; self.b = []; self.basicas = []
        for (a, s, b), extras in zip(self.restricciones, info_rest):
            fila = list(a) + [Fraction(0)] * (self.n_total - self.n_dec)
            basica_i = None
            for tp, k, _ in extras:
                if tp == "S+":  fila[k] = Fraction(1);  basica_i = k
                elif tp == "S-": fila[k] = Fraction(-1)
                elif tp == "A":  fila[k] = Fraction(1);  basica_i = k
            self.A.append(fila); self.b.append(b); self.basicas.append(basica_i)
        self._info_rest = info_rest

    # ---- Cálculos por iteración ----
    def _calcular_zj_y_diff(self):
        Zj = []
        for j in range(self.n_total):
            s = Coef(0)
            for i in range(len(self.A)):
                s = s + self.C[self.basicas[i]] * self.A[i][j]
            Zj.append(s)
        diff = [self.C[j] - Zj[j] for j in range(self.n_total)]
        Z = Coef(0)
        for i in range(len(self.A)):
            Z = Z + self.C[self.basicas[i]] * self.b[i]
        return Zj, diff, Z

    def _elegir_columna_pivote(self, diff):
        mejor = None
        if self.tipo == "max":
            for j, d in enumerate(diff):
                pos = (d.m > 0) or (d.m == 0 and d.const > 0)
                if pos and (mejor is None or d.mas_positivo_que(diff[mejor])):
                    mejor = j
        else:
            for j, d in enumerate(diff):
                neg = (d.m < 0) or (d.m == 0 and d.const < 0)
                if neg and (mejor is None or d.mas_negativo_que(diff[mejor])):
                    mejor = j
        return mejor

    def _elegir_fila_pivote(self, col):
        mejor_i = None; mejor_q = None
        for i in range(len(self.A)):
            aij = self.A[i][col]
            if aij <= 0: continue
            q = self.b[i] / aij
            if mejor_q is None or q < mejor_q:
                mejor_q = q; mejor_i = i
        return mejor_i, mejor_q

    def _calcular_cocientes(self, col):
        out = []
        for i in range(len(self.A)):
            aij = self.A[i][col]
            out.append("---" if aij <= 0 else fmt_num(self.b[i] / aij))
        return out

    def _snapshot(self, Zj, diff, Z, cocientes, col_piv, fila_piv, ops_prev, es_opt):
        """Captura el estado actual (deepcopy para no perderlo al pivotear)."""
        return {
            "A":       [list(f) for f in self.A],
            "b":       list(self.b),
            "basicas": list(self.basicas),
            "Cb":      [self.C[k] for k in self.basicas],
            "VB":      [self.nombres_var[k] for k in self.basicas],
            "Zj":      list(Zj),
            "Cj_Zj":   list(diff),
            "Z":       Z,
            "coc":     list(cocientes) if cocientes else None,
            "col_piv": col_piv,
            "fila_piv": fila_piv,
            "ops_prev": list(ops_prev) if ops_prev else None,
            "es_optima": es_opt,
        }

    def _pivotear(self, fila_piv, col_piv):
        ops = []
        piv = self.A[fila_piv][col_piv]
        n_new = self.nombres_var[col_piv]
        if piv != 1:
            inv = Fraction(1) / piv
            self.A[fila_piv] = [a / piv for a in self.A[fila_piv]]
            self.b[fila_piv] = self.b[fila_piv] / piv
            ops.append(f"F{fila_piv+1}_nueva = ({fmt_num(inv)}) F{fila_piv+1}"
                       f"   [normalizar el pivote: {n_new} entra con coef 1]")
        else:
            ops.append(f"F{fila_piv+1}_nueva = F{fila_piv+1}   [pivote ya es 1]")
        for i in range(len(self.A)):
            if i == fila_piv: continue
            factor = self.A[i][col_piv]
            if factor == 0: continue
            self.A[i] = [self.A[i][j] - factor * self.A[fila_piv][j]
                         for j in range(self.n_total)]
            self.b[i] = self.b[i] - factor * self.b[fila_piv]
            if factor > 0:
                ops.append(f"F{i+1}_nueva = F{i+1} - ({fmt_num(factor)}) F{fila_piv+1}")
            else:
                ops.append(f"F{i+1}_nueva = F{i+1} + ({fmt_num(-factor)}) F{fila_piv+1}")
        self.basicas[fila_piv] = col_piv
        return ops

    def resolver(self):
        """Devuelve un dict-registro con todo el procedimiento."""
        record = {
            "tipo":        self.tipo,
            "c_orig":      list(self.c_orig),
            "n_dec":       self.n_dec,
            "nombres_dec": list(self.nombres_dec),
            "restricciones_orig": [(list(a), s, b) for a, s, b in self.restricciones],
            "C":           list(self.C),
            "nombres_var": list(self.nombres_var),
            "tipo_var":    list(self.tipo_var),
            "n_total":     self.n_total,
            "iteraciones": [],
            "estado":      None,   # 'optima' | 'no_acotada' | 'infactible' | 'limite'
            "Z_final":     None,
            "valores":     None,
        }

        ops_prev = None
        for it in range(self.max_iter + 1):
            Zj, diff, Z = self._calcular_zj_y_diff()
            col_piv = self._elegir_columna_pivote(diff)

            if col_piv is None:
                snap = self._snapshot(Zj, diff, Z, None, None, None, ops_prev, True)
                record["iteraciones"].append(snap)
                # Verificar infactibilidad
                for i, k in enumerate(self.basicas):
                    if self.tipo_var[k] == "A" and self.b[i] > 0:
                        record["estado"] = "infactible"
                        record["Z_final"] = Z
                        return record
                record["estado"] = "optima"
                record["Z_final"] = Z
                valores = {n: Fraction(0) for n in self.nombres_var}
                for i, k in enumerate(self.basicas):
                    valores[self.nombres_var[k]] = self.b[i]
                record["valores"] = valores
                return record

            cocientes = self._calcular_cocientes(col_piv)
            fila_piv, _ = self._elegir_fila_pivote(col_piv)

            snap = self._snapshot(Zj, diff, Z, cocientes, col_piv, fila_piv, ops_prev, False)
            record["iteraciones"].append(snap)

            if fila_piv is None:
                record["estado"] = "no_acotada"
                return record

            ops_prev = self._pivotear(fila_piv, col_piv)

        record["estado"] = "limite"
        return record


# ============================================================================
#  PARSER de lenguaje matemático natural
# ============================================================================
def _norm_text(s):
    s = s.replace("≤", "<=").replace("≥", ">=")
    s = s.replace("−", "-").replace("–", "-").replace("—", "-")
    s = s.replace("·", "*")
    return s

def _parse_number(s):
    s = s.strip().replace(",", ".")
    if "/" in s:
        a, b = s.split("/", 1)
        return Fraction(int(a.strip()), int(b.strip()))
    if "." in s:
        return Fraction(s)
    return Fraction(int(s))

def _norm_var(s):
    """X_1, x1, X1 -> X1"""
    return s.upper().replace("_", "")

def _tokenize(expr):
    tokens = []; i = 0; n = len(expr)
    while i < n:
        c = expr[i]
        if c.isspace() or c == '*':
            i += 1
        elif c in '+-':
            tokens.append(('op', c)); i += 1
        elif c.isdigit() or c == '.':
            j = i
            while j < n and (expr[j].isdigit() or expr[j] in '.,/'):
                j += 1
            tokens.append(('num', expr[i:j])); i = j
        elif c.isalpha() or c == '_':
            j = i
            while j < n and (expr[j].isalnum() or expr[j] == '_'):
                j += 1
            tokens.append(('var', expr[i:j])); i = j
        else:
            i += 1
    return tokens

def _parse_linear(expr):
    """Devuelve dict {var: coef} de una expresión lineal."""
    tokens = _tokenize(expr)
    coefs = {}; sign = 1; i = 0; n = len(tokens)
    while i < n:
        kind, val = tokens[i]
        if kind == 'op':
            sign = 1 if val == '+' else -1
            i += 1; continue
        coef = None; var = None
        if kind == 'num':
            coef = _parse_number(val); i += 1
            if i < n and tokens[i][0] == 'var':
                var = tokens[i][1]; i += 1
        elif kind == 'var':
            var = val; coef = Fraction(1); i += 1
        else:
            i += 1; continue
        if coef is None: coef = Fraction(1)
        coef = coef * sign
        if var:
            v = _norm_var(var)
            coefs[v] = coefs.get(v, Fraction(0)) + coef
        sign = 1
    return coefs


def parse_problem(text):
    """
    Parse natural-language LP problem.
    Devuelve: (tipo, c, restricciones, var_names)
        tipo: 'max' o 'min'
        c: lista de Fractions (coef FO en orden de var_names)
        restricciones: lista de (a_list, sign, b_Fraction)
        var_names: lista de variables ordenadas
    """
    text = _norm_text(text)
    lines = [l.strip() for l in text.split("\n")]
    lines = [l for l in lines if l]

    tipo = None
    fo_expr = None
    raw_rests = []   # lista de (coefs_dict, sign, rhs)
    skip_markers = ("s.a", "sujeto a", "subject to", "s.t", "restricciones",
                    "no negat", "negat")

    for line in lines:
        low = line.lower()
        # Líneas marcador que solo dicen "s.a." o similar
        if line.strip().lower().rstrip(":.") in ("s.a", "sujeto a", "s.t", "subject to",
                                                  "restricciones"):
            continue

        # Detectar Max/Min
        if tipo is None:
            m = re.search(r'\b(max|min|maximizar|minimizar|maximize|minimize)\w*\b',
                          low)
            if m:
                tipo = 'max' if m.group(1).startswith(('max', 'maxim')) else 'min'
                # Quitar la palabra Max/Min y "Z =" si existe
                rest = re.sub(r'\b(max|min|maximizar|minimizar|maximize|minimize)\w*\.?\s*',
                              '', line, count=1, flags=re.IGNORECASE)
                rest = re.sub(r'^\s*z\s*=\s*', '', rest, flags=re.IGNORECASE)
                fo_expr = rest.strip()
                # Si la línea solo tenía "Max" o "Maximizar", la FO está en la siguiente
                # línea — entonces dejar fo_expr vacío y procesar siguientes líneas
                continue

        # Si tipo ya está pero fo_expr está vacío, esta línea es la FO
        if tipo is not None and not fo_expr:
            # Pero solo si no tiene operador de comparación
            if not re.search(r'(<=|>=|=)', line):
                rest = re.sub(r'^\s*z\s*=\s*', '', line, flags=re.IGNORECASE)
                fo_expr = rest.strip()
                continue

        # Detectar restricción (contiene <=, >=, = pero no es definición de Z)
        m = re.search(r'(<=|>=|=)', line)
        if m:
            sign = m.group(1)
            lhs, rhs = line.split(sign, 1)

            # Saltar "Z = ..."
            if re.search(r'\bz\b', lhs, re.IGNORECASE) and sign == "=":
                # Es la FO escrita como "Z = ..."
                if tipo is None:
                    continue  # ignorar si no hay tipo aún
                fo_expr = rhs.strip()
                continue

            # Saltar no-negatividad: "X1, X2, ... >= 0" o "X1 ≥ 0" con coma
            rhs_clean = rhs.strip()
            try:
                rhs_val = _parse_number(rhs_clean)
            except (ValueError, ZeroDivisionError):
                continue
            if sign == ">=" and rhs_val == 0:
                # ¿lhs solo contiene variables y comas?
                if re.match(r'^[\sA-Za-z0-9_,]+$', lhs) and ',' in lhs:
                    continue
                # También saltamos si la línea menciona "no negativ"
                if 'negat' in low:
                    continue

            coefs = _parse_linear(lhs)
            if not coefs:
                continue
            raw_rests.append((coefs, sign, rhs_val))

    if tipo is None:
        raise ValueError("No se encontró 'Max' / 'Min' en el modelo.")
    if not fo_expr:
        raise ValueError("No se encontró la función objetivo.")
    if not raw_rests:
        raise ValueError("No se encontraron restricciones.")

    fo_coefs = _parse_linear(fo_expr)
    if not fo_coefs:
        raise ValueError("No se pudo parsear la función objetivo.")

    # Orden de variables: primero las que aparecen en FO, luego nuevas en restricciones
    var_names = list(fo_coefs.keys())
    for coefs, _, _ in raw_rests:
        for v in coefs:
            if v not in var_names:
                var_names.append(v)

    # Ordenar X1, X2, ... numéricamente
    def sort_key(v):
        m = re.match(r'^([A-Z]+)(\d+)$', v)
        if m: return (m.group(1), int(m.group(2)))
        return (v, 0)
    var_names = sorted(var_names, key=sort_key)

    c = [fo_coefs.get(v, Fraction(0)) for v in var_names]
    restricciones = []
    for coefs, sign, rhs in raw_rests:
        a = [coefs.get(v, Fraction(0)) for v in var_names]
        restricciones.append((a, sign, rhs))

    return tipo, c, restricciones, var_names


# ============================================================================
#  PARSER CON IA (Groq - modelo GPT OSS 120B) - para enunciados en lenguaje natural
# ============================================================================
PROMPT_SYSTEM_IA = """Eres un experto en Investigación de Operaciones. Tu UNICA tarea es extraer el modelo de Programación Lineal (PL) a partir de un enunciado de problema en español.

REGLAS ESTRICTAS:
1. Responde SOLO con un objeto JSON valido. SIN texto antes ni despues. SIN bloques de codigo markdown.
2. Identifica las variables de decision. Nómbralas X1, X2, ..., Xn en orden logico (segun aparecen en el enunciado).
3. La funcion objetivo: identifica si es 'max' (maximizar) o 'min' (minimizar) y los coeficientes (uno por variable, en el mismo orden).
4. Las restricciones: identifica TODAS las restricciones explicitas del enunciado (recursos limitados, demandas, ratios, etc.).
5. NO incluyas la restriccion de no-negatividad (X >= 0) en la lista.
6. Los coeficientes deben ser numeros (enteros o decimales). Si el enunciado menciona fracciones como '1/2', usa 0.5.
7. Si el enunciado dice 'ingreso bruto' o 'ingreso', usa el PRECIO directamente. Si dice 'beneficio', 'utilidad' o 'ganancia', usa (precio - costo). Si dice 'costo', usa el costo.
8. Sé fiel al enunciado. NO inventes restricciones que no esten explicitas.
9. Para cada restriccion, identifica correctamente el signo: 'no puede exceder X', 'maximo de X', 'limitado a X' = '<='; 'al menos X', 'minimo de X', 'no menor que X' = '>='; 'igual a X' = '='.
10. Cuidado con frases como 'no puede ser mayor que la otra en más de N': eso es 'A - B <= N' (NO 'B - A <= N').

FORMATO DE RESPUESTA (JSON exacto):
{
  "tipo": "max" o "min",
  "variables": ["X1", "X2", ...],
  "descripcion_variables": {"X1": "descripcion breve", "X2": "..."},
  "fo_coeficientes": [num, num, ...],
  "fo_descripcion": "que representa Z (ej: ingreso, beneficio, costo)",
  "restricciones": [
    {"coeficientes": [num, num, ...], "signo": "<=" o ">=" o "=", "rhs": num, "descripcion": "breve"}
  ],
  "advertencias": ["..."]
}"""


def parse_enunciado_ia(texto, api_key, model="llama-3.3-70b-versatile"):
    """Usa la API de Groq para extraer el modelo de PL a partir de un enunciado
    en lenguaje natural.

    Devuelve: (tipo, c, restricciones, var_names, metadatos)
    Lanza ValueError si la IA no devuelve un JSON valido.
    """
    if not api_key:
        raise ValueError("No se proporciono API key de Groq.")

    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": PROMPT_SYSTEM_IA},
            {"role": "user",
             "content": ("Extrae el modelo de PL del siguiente enunciado y responde "
                         "UNICAMENTE con el objeto JSON pedido (sin texto antes ni despues, "
                         "sin bloques markdown):\n\n" + texto)},
        ],
        "temperature": 0,
        "max_tokens": 2000,
    }

    body = _json.dumps(payload).encode("utf-8")

    req = urllib.request.Request(
        "https://api.groq.com/openai/v1/chat/completions",
        data=body,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": "SimplexGUI/1.0 (Python; tkinter; educational use)",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            raw = resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        try:
            err_body = e.read().decode("utf-8")
            err_data = _json.loads(err_body)
            msg = err_data.get("error", {}).get("message", err_body)
        except Exception:
            msg = str(e)
        if e.code == 401:
            raise ValueError(f"API key invalida o no autorizada. Detalle: {msg}")
        if e.code == 403:
            raise ValueError(
                f"Modelo no disponible o sin permiso. Prueba con otro modelo "
                f"(p.ej. llama-3.3-70b-versatile, llama-3.1-8b-instant, "
                f"meta-llama/llama-4-scout-17b-16e-instruct) desde el boton "
                f"'Configurar API key'.\n\nDetalle de Groq: {msg}")
        if e.code == 413:
            raise ValueError(
                f"Payload demasiado grande. Acorta el enunciado o usa otro modelo.\n"
                f"Detalle: {msg}")
        if e.code == 429:
            raise ValueError(f"Limite de uso alcanzado (rate limit). Espera unos segundos.\nDetalle: {msg}")
        if e.code == 400:
            raise ValueError(
                f"Peticion mal formada (HTTP 400). Probablemente el modelo "
                f"'{model}' no acepta algun parametro. Detalle: {msg}")
        raise ValueError(f"Error HTTP {e.code} de Groq: {msg}")
    except urllib.error.URLError as e:
        raise ValueError(f"No se pudo conectar con Groq. Verifica tu conexion a internet.\nDetalle: {e}")

    data = _json.loads(raw)
    if "choices" not in data or not data["choices"]:
        raise ValueError(f"Respuesta inesperada de Groq: {raw[:500]}")
    full = data["choices"][0]["message"]["content"].strip()

    # Quitar bloques markdown si el modelo los puso
    if full.startswith("```"):
        lines = full.split("\n")
        if lines[-1].strip().startswith("```"):
            full = "\n".join(lines[1:-1])
        else:
            full = "\n".join(lines[1:])
        full = full.strip()

    # Si el modelo agrego texto antes/despues del JSON, intentar extraer el {...}
    if not full.startswith("{"):
        m_first = full.find("{")
        m_last  = full.rfind("}")
        if m_first != -1 and m_last != -1 and m_last > m_first:
            full = full[m_first:m_last + 1]

    # Parsear JSON
    try:
        data = _json.loads(full)
    except _json.JSONDecodeError as e:
        raise ValueError(f"La IA no devolvio JSON valido. Detalle: {e}\n\nRespuesta:\n{full[:500]}")

    # Validar estructura mínima
    for campo in ("tipo", "variables", "fo_coeficientes", "restricciones"):
        if campo not in data:
            raise ValueError(f"Falta el campo '{campo}' en la respuesta de la IA.")
    if data["tipo"] not in ("max", "min"):
        raise ValueError(f"Tipo invalido: '{data['tipo']}' (debe ser max o min).")
    if len(data["variables"]) != len(data["fo_coeficientes"]):
        raise ValueError("El nro. de variables no coincide con los coeficientes de la FO.")

    tipo = data["tipo"]
    var_names = [v.upper().replace("_", "") for v in data["variables"]]
    c = [_num_a_frac(x) for x in data["fo_coeficientes"]]

    restricciones = []
    for r in data["restricciones"]:
        if "coeficientes" not in r or "signo" not in r or "rhs" not in r:
            raise ValueError(f"Restriccion mal formada: {r}")
        if r["signo"] not in ("<=", ">=", "="):
            raise ValueError(f"Signo invalido en restriccion: '{r['signo']}'.")
        if len(r["coeficientes"]) != len(var_names):
            raise ValueError(
                f"La restriccion tiene {len(r['coeficientes'])} coef. pero hay {len(var_names)} variables.")
        a = [_num_a_frac(x) for x in r["coeficientes"]]
        restricciones.append((a, r["signo"], _num_a_frac(r["rhs"])))

    metadatos = {
        "descripcion_variables": data.get("descripcion_variables", {}),
        "fo_descripcion": data.get("fo_descripcion", ""),
        "advertencias": data.get("advertencias", []),
        "restricciones_desc": [r.get("descripcion", "") for r in data["restricciones"]],
        "modelo_ia": model,
    }

    return tipo, c, restricciones, var_names, metadatos


def _num_a_frac(x):
    """Convierte un valor (int, float, str) a Fraction de forma exacta."""
    if isinstance(x, Fraction):
        return x
    if isinstance(x, int):
        return Fraction(x)
    if isinstance(x, float):
        # Limitar denominador para evitar fracciones gigantes por imprecisión
        return Fraction(x).limit_denominator(100000)
    if isinstance(x, str):
        s = x.strip().replace(",", ".")
        if "/" in s:
            a, b = s.split("/", 1)
            return Fraction(int(a.strip()), int(b.strip()))
        if "." in s:
            return Fraction(s).limit_denominator(100000)
        return Fraction(int(s))
    raise ValueError(f"Tipo numerico no soportado: {type(x)}")


def es_enunciado_natural(texto):
    """Heuristica: decide si el texto es un enunciado en lenguaje natural
    o un modelo matematico formal."""
    t = texto.lower()
    # Si tiene Max/Min seguido de Z = o coeficientes con variables, es matematico
    if re.search(r'\b(max|min|maximizar|minimizar)\w*\s*z?\s*=?\s*[-+]?\s*\d', t):
        return False
    # Si tiene varias lineas con <= o >=, es matematico
    n_restr = len(re.findall(r'(<=|>=|≤|≥)', texto))
    if n_restr >= 2:
        return False
    # Si tiene muchas palabras de prosa (mas de 30 palabras), es natural
    palabras = re.findall(r'\w+', texto)
    if len(palabras) > 30:
        return True
    return False
def _str_combinacion(coefs, nombres, incluir_ceros=False):
    partes = []; primero = True
    for c, n in zip(coefs, nombres):
        if isinstance(c, Coef):
            if c.es_cero() and not incluir_ceros: continue
            if primero:
                partes.append(n if (c.const == 1 and c.m == 0) else f"{c} {n}")
                primero = False
            else:
                if c.m == 0 and c.const < 0:
                    partes.append(f"- {n}" if c.const == -1 else f"- {fmt_num(-c.const)} {n}")
                elif c.m == 0:
                    partes.append(f"+ {n}" if c.const == 1 else f"+ {fmt_num(c.const)} {n}")
                else:
                    partes.append(f"+ ({c}) {n}")
        else:
            if c == 0 and not incluir_ceros: continue
            if primero:
                partes.append(n if c == 1 else (f"-{n}" if c == -1 else f"{fmt_num(c)} {n}"))
                primero = False
            else:
                if c == 1: partes.append(f"+ {n}")
                elif c == -1: partes.append(f"- {n}")
                elif c < 0: partes.append(f"- {fmt_num(-c)} {n}")
                else:       partes.append(f"+ {fmt_num(c)} {n}")
    return " ".join(partes) if partes else "0"


def format_record_text(record, metadatos=None):
    """Convierte el registro completo en un string para mostrar."""
    out = []
    ANCHO = 78
    SEP   = "=" * ANCHO

    # ---------- Encabezado ----------
    verbo_max = "MAXIMIZACION" if record["tipo"] == "max" else "MINIMIZACION"
    verbo_short = "Max" if record["tipo"] == "max" else "Min"

    out.append(SEP)
    out.append(f"  RESOLUCION POR EL METODO SIMPLEX (Gran M) - {verbo_max}")
    out.append(SEP)

    # ---------- Modelo general ----------
    out.append("")
    out.append("-" * ANCHO)
    out.append(" MODELO GENERAL")
    out.append("-" * ANCHO)
    fo = _str_combinacion(record["c_orig"], record["nombres_dec"])
    out.append(f"  F.O :  {verbo_short}.  Z = {fo}")
    out.append("  s.a (sujeto a):")
    for i, (a, s, b) in enumerate(record["restricciones_orig"], 1):
        izq = _str_combinacion(a, record["nombres_dec"])
        out.append(f"    {i}) {izq} {s} {fmt_num(b)}")
    out.append(f"  {', '.join(record['nombres_dec'])} >= 0")

    # ---------- Modelo estándar ----------
    out.append("")
    out.append("-" * ANCHO)
    out.append(" MODELO ESTANDAR (forma canonica del Simplex)")
    out.append("-" * ANCHO)
    out.append("  Reglas usadas:")
    out.append("    - Cada '<='  recibe HOLGURA  S (coef +1).")
    out.append("    - Cada '>='  recibe EXCEDENTE S (coef -1) + ARTIFICIAL A (coef +1).")
    out.append("    - Cada '='   recibe ARTIFICIAL A.")
    out.append("    - En la F.O., las A llevan -M (Max) o +M (Min).")
    out.append("")
    fo_e = _str_combinacion(record["C"], record["nombres_var"], incluir_ceros=True)
    out.append(f"  F.O :  {verbo_short}.  Z = {fo_e}")
    out.append("  s.a:")
    snap0 = record["iteraciones"][0]
    for i, (fila, b) in enumerate(zip(snap0["A"], snap0["b"]), 1):
        izq = _str_combinacion(fila, record["nombres_var"])
        out.append(f"    {i}) {izq} = {fmt_num(b)}")

    # Criterios
    out.append("")
    out.append("-" * ANCHO)
    out.append(" CRITERIOS DEL METODO")
    out.append("-" * ANCHO)
    if record["tipo"] == "max":
        out.append("  - VARIABLE QUE ENTRA : la del MAYOR Cj-Zj POSITIVO.")
        out.append("  - OPTIMO            : todos los Cj-Zj son <= 0.")
    else:
        out.append("  - VARIABLE QUE ENTRA : la del MAYOR Cj-Zj NEGATIVO.")
        out.append("  - OPTIMO            : todos los Cj-Zj son >= 0.")
    out.append("  - VARIABLE QUE SALE  : la del MENOR cociente POSITIVO b/aij.")
    out.append("  - PIVOTEO            : F_piv <- F_piv / piv ; F_otra <- F_otra - aij * F_piv.")

    # ---------- Iteraciones ----------
    for it_idx, snap in enumerate(record["iteraciones"]):
        out.append("")
        if snap["es_optima"]:
            out.append("+" + "-" * (ANCHO - 2) + "+")
            out.append(f"| ITERACION {it_idx}  --  TABLA OPTIMA"
                       + " " * (ANCHO - 38) + "|")
            out.append("+" + "-" * (ANCHO - 2) + "+")
        else:
            out.append("+" + "-" * (ANCHO - 2) + "+")
            out.append(f"| ITERACION {it_idx}" + " " * (ANCHO - 16) + "|")
            out.append("+" + "-" * (ANCHO - 2) + "+")

        if snap["ops_prev"]:
            out.append("  Operaciones de fila aplicadas:")
            for op in snap["ops_prev"]:
                out.append(f"    {op}")
            out.append("")

        out.append(_render_tabla_texto(record, snap))

        if not snap["es_optima"]:
            ent = record["nombres_var"][snap["col_piv"]]
            sal = snap["VB"][snap["fila_piv"]]
            piv = snap["A"][snap["fila_piv"]][snap["col_piv"]]
            dz  = snap["Cj_Zj"][snap["col_piv"]]
            tag = "MAYOR positivo" if record["tipo"] == "max" else "MAYOR negativo"
            out.append("")
            out.append(f"  >>> Variable que ENTRA: {ent}  (Cj-Zj = {dz}, el {tag})")
            out.append(f"  >>> Variable que SALE : {sal}  (menor cociente positivo b/aij)")
            out.append(f"  >>> Pivote            : a[F{snap['fila_piv']+1}][{ent}]"
                       f" = {fmt_num(piv)}  (marcado entre [corchetes])")

    # ---------- Solución ----------
    out.append("")
    out.append(SEP)
    if record["estado"] == "optima":
        out.append("  SOLUCION OPTIMA ENCONTRADA")
        out.append(SEP)

        valores    = record["valores"]
        nombres_dec = record["nombres_dec"]
        desc_vars  = {}
        fo_desc    = ""
        if metadatos:
            desc_vars = metadatos.get("descripcion_variables", {})
            fo_desc   = metadatos.get("fo_descripcion", "")

        SUB = "  " + "-" * (ANCHO - 4)

        # --- [1] Identificacion de variables ---
        if desc_vars:
            out.append("")
            out.append("  [1] IDENTIFICACION DE VARIABLES DE DECISION")
            out.append(SUB)
            for n in nombres_dec:
                desc = desc_vars.get(n, "")
                out.append(f"  {n:<6} =  {desc}" if desc else f"  {n}")

        # --- [2] Valores optimos de las variables de decision ---
        out.append("")
        out.append("  [2] VALORES OPTIMOS DE LAS VARIABLES DE DECISION")
        out.append(SUB)
        for n in nombres_dec:
            v    = valores[n]
            ap   = f"  (~{float(v):.4f})" if (isinstance(v, Fraction) and v.denominator != 1) else ""
            desc = desc_vars.get(n, "")
            tag  = f"   [{desc}]" if desc else ""
            out.append(f"  {n} = {fmt_num(v)}{ap}{tag}")

        # --- [3] Analisis de holguras y excedentes ---
        out.append("")
        out.append("  [3] ANALISIS DE HOLGURAS / EXCEDENTES")
        out.append(SUB)
        alguna = False
        for j in range(record["n_dec"], record["n_total"]):
            n  = record["nombres_var"][j]
            tp = record["tipo_var"][j]
            v  = valores[n]
            if tp == "A":
                continue
            alguna = True
            num_rest = re.sub(r'[^0-9]', '', n)
            r_label  = f"R{num_rest}" if num_rest else "?"
            ap = f"  (~{float(v):.4f})" if (isinstance(v, Fraction) and v.denominator != 1) else ""
            if tp == "S+":
                if v == 0:
                    out.append(f"  {n} = 0   -> Restriccion {r_label}: ACTIVA (recurso agotado, sin holgura)")
                else:
                    out.append(f"  {n} = {fmt_num(v)}{ap}   -> Restriccion {r_label}: holgura de {fmt_num(v)} unidades (recurso no agotado)")
            elif tp == "S-":
                if v == 0:
                    out.append(f"  {n} = 0   -> Restriccion {r_label}: ACTIVA (se cumple exactamente el minimo exigido)")
                else:
                    out.append(f"  {n} = {fmt_num(v)}{ap}   -> Restriccion {r_label}: excedente de {fmt_num(v)} unidades sobre el minimo")
        if not alguna:
            out.append("  (No hay variables de holgura/excedente en este modelo)")

        # --- [4] Verificacion de restricciones ---
        out.append("")
        out.append("  [4] VERIFICACION DE RESTRICCIONES (valores optimos sustituidos)")
        out.append(SUB)
        for i, (a, s, b) in enumerate(record["restricciones_orig"], 1):
            lhs_val = sum(Fraction(a[j]) * valores[nombres_dec[j]] for j in range(len(nombres_dec)))
            terms = []
            for j, (coef, nom) in enumerate(zip(a, nombres_dec)):
                cf = Fraction(coef)
                vj = valores[nom]
                if cf != 0:
                    terms.append(f"({fmt_num(cf)})({fmt_num(vj)})")
            lhs_str = " + ".join(terms) if terms else "0"
            ok = ((s == "<=" and lhs_val <= b) or
                  (s == ">=" and lhs_val >= b) or
                  (s == "="  and lhs_val == b))
            satisf = "SATISFECHA" if ok else "NO SATISFECHA"
            out.append(f"  R{i}: {lhs_str} = {fmt_num(lhs_val)}  {s}  {fmt_num(b)}   [{satisf}]")

        # --- [5] Valor de la funcion objetivo ---
        out.append("")
        out.append("  [5] VALOR DE LA FUNCION OBJETIVO")
        out.append(SUB)
        Z = record["Z_final"]
        ap = ""
        if isinstance(Z, Coef) and Z.m == 0 and Z.const.denominator != 1:
            ap = f"  (~{float(Z.const):.4f})"
        verbo_obj = "MAXIMO" if record["tipo"] == "max" else "MINIMO"
        fo_tag = f"  ({fo_desc})" if fo_desc else ""
        out.append(f"  Z = {Z}{ap}   <- Valor {verbo_obj} de la F.O.{fo_tag}")

        # --- [6] Conclusion / interpretacion ---
        out.append("")
        out.append("  [6] CONCLUSION / INTERPRETACION")
        out.append(SUB)
        accion = "MAXIMIZAR" if record["tipo"] == "max" else "MINIMIZAR"
        out.append(f"  Para {accion} la funcion objetivo Z, la solucion optima indica:")
        out.append("")
        for n in nombres_dec:
            v    = valores[n]
            desc = desc_vars.get(n, "")
            ap2  = f"  (~{float(v):.4f})" if (isinstance(v, Fraction) and v.denominator != 1) else ""
            tag  = f"  ->  {desc}" if desc else ""
            out.append(f"    * {n} = {fmt_num(v)}{ap2}{tag}")
        out.append("")
        activas     = []
        con_holgura = []
        for j in range(record["n_dec"], record["n_total"]):
            nv  = record["nombres_var"][j]
            tp  = record["tipo_var"][j]
            vv  = valores[nv]
            if tp not in ("S+", "S-"):
                continue
            num_rest = re.sub(r'[^0-9]', '', nv)
            if not num_rest:
                continue
            idx_r = int(num_rest)
            if vv == 0:
                activas.append(idx_r)
            else:
                tipo_h = "holgura" if tp == "S+" else "excedente"
                con_holgura.append((idx_r, fmt_num(vv), tipo_h))
        if activas:
            r_str = ", ".join(f"R{r_}" for r_ in sorted(activas))
            out.append(f"  Las restricciones {r_str} son ACTIVAS (se usan al 100% de su capacidad).")
        for idx_r, val, tipo_h in sorted(con_holgura):
            out.append(f"  La restriccion R{idx_r} tiene {tipo_h} de {val} unidades (capacidad no utilizada del todo).")
        verbo_obj2 = "maximo" if record["tipo"] == "max" else "minimo"
        out.append("")
        out.append(f"  El valor {verbo_obj2} de la funcion objetivo es  Z = {Z}.")
        out.append(SEP)

    elif record["estado"] == "infactible":
        out.append(" RESULTADO: PROBLEMA INFACTIBLE")
        out.append(SEP)
        out.append("  Una variable artificial quedó en la base con valor > 0.")
        out.append("  Las restricciones no tienen solución factible.")
    elif record["estado"] == "no_acotada":
        out.append(" RESULTADO: PROBLEMA NO ACOTADO")
        out.append(SEP)
        out.append("  La columna pivote no tiene ningún coeficiente positivo.")
        out.append("  La función objetivo puede crecer/decrecer sin límite.")
    else:
        out.append(" LIMITE DE ITERACIONES ALCANZADO")
        out.append(SEP)

    return "\n".join(out)


def _render_tabla_texto(record, snap):
    """Renderiza una tabla simplex en formato texto (estilo del profesor)."""
    n_var  = record["n_total"]
    n_rest = len(snap["A"])
    nombres = record["nombres_var"]
    C = record["C"]

    # Cabecera 1: Cj
    fila_cj  = ["", "", "Cj->"] + [str(C[j]) for j in range(n_var)] + ["", ""]
    # Cabecera 2: nombres
    fila_var = ["Fila", "Cb", "V.B."] + list(nombres) + ["b", "b/aij"]

    filas_dat = []
    for i in range(n_rest):
        f = [f"F{i+1}", str(snap["Cb"][i]), snap["VB"][i]]
        for j in range(n_var):
            f.append(fmt_num(snap["A"][i][j]))
        f.append(fmt_num(snap["b"][i]))
        f.append(snap["coc"][i] if snap["coc"] else "")
        filas_dat.append(f)

    # Marcar pivote
    if snap["col_piv"] is not None and snap["fila_piv"] is not None:
        cp = snap["col_piv"]; rp = snap["fila_piv"]
        filas_dat[rp][3 + cp] = f"[{filas_dat[rp][3 + cp]}]"
        filas_dat[rp][0]      = f">{filas_dat[rp][0]}"

    fila_zj   = ["", "", "Zj"]   + [str(snap["Zj"][j])   for j in range(n_var)] + [str(snap["Z"]), ""]
    fila_diff = ["", "", "Cj-Zj"]+ [str(snap["Cj_Zj"][j]) for j in range(n_var)] + ["", ""]

    todas = [fila_cj, fila_var] + filas_dat + [fila_zj, fila_diff]
    ncols = len(fila_var)
    anchos = [0] * ncols
    for fila in todas:
        for k in range(ncols):
            anchos[k] = max(anchos[k], len(str(fila[k])))

    def render(fila):
        return " | ".join(str(x).rjust(anchos[k]) for k, x in enumerate(fila))
    sep = "=+=".join("=" * a for a in anchos)

    out_lines = []
    out_lines.append(render(fila_cj))
    out_lines.append(render(fila_var))
    out_lines.append(sep)
    for f in filas_dat: out_lines.append(render(f))
    out_lines.append(sep)
    out_lines.append(render(fila_zj))
    out_lines.append(render(fila_diff))
    return "\n".join(out_lines)


# ============================================================================
#  EXPORTADOR A EXCEL  (formato del profesor con colores)
# ============================================================================
class ExcelExporter:

    def __init__(self):
        if not OPENPYXL_OK:
            raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")

        # Colores estilo profesor (definidos aquí para no fallar si falta openpyxl)
        self.F_ORANGE   = PatternFill('solid', fgColor='FFC000')
        self.F_RED      = PatternFill('solid', fgColor='FF0000')
        self.F_LIGHTORG = PatternFill('solid', fgColor='FCE4A6')
        self.F_HEADER   = PatternFill('solid', fgColor='D9D9D9')
        self.F_GREEN    = PatternFill('solid', fgColor='C6EFCE')
        self.F_TITLE    = PatternFill('solid', fgColor='305496')

        self.THIN_S    = Side(style='thin', color='808080')
        self.MEDIUM_S  = Side(style='medium', color='000000')
        self.DIAG_S    = Side(style='thin', color='000000')

        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.border_all = Border(left=self.THIN_S, right=self.THIN_S,
                                 top=self.THIN_S, bottom=self.THIN_S)
        self.border_med = Border(left=self.MEDIUM_S, right=self.MEDIUM_S,
                                 top=self.MEDIUM_S, bottom=self.MEDIUM_S)
        self.font_bold  = Font(bold=True)
        self.font_white_bold = Font(bold=True, color='FFFFFF')
        self.font_title = Font(bold=True, size=14, color='FFFFFF')
        self.center     = Alignment(horizontal='center', vertical='center')
        self.left_wrap  = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # ----------------------------------------------------------------
    def export(self, record, enunciado="", metadatos=None, filename="simplex.xlsx"):
        self._add_enunciado(enunciado)
        self._add_modelo(record)
        for idx, snap in enumerate(record["iteraciones"]):
            self._add_iteracion(record, idx, snap)
        self._add_solucion(record, metadatos)
        self.wb.save(filename)

    # ----------------------------------------------------------------
    def _title_row(self, ws, row, text, ncols=14):
        ws.cell(row=row, column=1, value=text)
        ws.cell(row=row, column=1).font = self.font_title
        ws.cell(row=row, column=1).fill = self.F_TITLE
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        ws.row_dimensions[row].height = 24

    def _add_enunciado(self, texto):
        ws = self.wb.create_sheet("Enunciado")
        self._title_row(ws, 1, "ENUNCIADO DEL PROBLEMA", ncols=10)
        ws.cell(row=3, column=1, value=texto or "(sin enunciado)")
        ws.cell(row=3, column=1).alignment = self.left_wrap
        ws.merge_cells(start_row=3, start_column=1, end_row=40, end_column=10)
        ws.column_dimensions['A'].width = 20
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 14

    def _add_modelo(self, record):
        ws = self.wb.create_sheet("Modelo Matematico")
        ncols = 12
        self._title_row(ws, 1, "MODELO MATEMATICO", ncols=ncols)

        # Modelo general
        r = 3
        ws.cell(row=r, column=1, value="MODELO GENERAL").font = Font(bold=True, size=12)
        ws.cell(row=r, column=1).fill = self.F_HEADER
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        r += 1
        verbo = "Max" if record["tipo"] == "max" else "Min"
        fo = _str_combinacion(record["c_orig"], record["nombres_dec"])
        ws.cell(row=r, column=1, value=f"F.O :  {verbo}.  Z = {fo}")
        ws.cell(row=r, column=1).font = self.font_bold
        r += 1
        ws.cell(row=r, column=1, value="s.a (sujeto a):")
        r += 1
        for i, (a, s, b) in enumerate(record["restricciones_orig"], 1):
            izq = _str_combinacion(a, record["nombres_dec"])
            ws.cell(row=r, column=1, value=f"  {i}) {izq} {s} {fmt_num(b)}")
            r += 1
        ws.cell(row=r, column=1,
                value=f"  {', '.join(record['nombres_dec'])} >= 0")
        r += 2

        # Modelo estándar
        ws.cell(row=r, column=1, value="MODELO ESTANDAR (forma canonica)").font = Font(bold=True, size=12)
        ws.cell(row=r, column=1).fill = self.F_HEADER
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        r += 1
        ws.cell(row=r, column=1, value="Reglas usadas:")
        r += 1
        reglas = [
            "  - Cada '<='  -> +S  (HOLGURA, coef +1)",
            "  - Cada '>='  -> -S + A  (EXCEDENTE y ARTIFICIAL)",
            "  - Cada '='   -> +A  (ARTIFICIAL)",
            "  - En F.O.: S llevan 0; A llevan -M (Max) o +M (Min).",
        ]
        for re_ in reglas:
            ws.cell(row=r, column=1, value=re_); r += 1
        r += 1
        fo_e = _str_combinacion(record["C"], record["nombres_var"], incluir_ceros=True)
        ws.cell(row=r, column=1, value=f"F.O :  {verbo}.  Z = {fo_e}")
        ws.cell(row=r, column=1).font = self.font_bold
        r += 1
        ws.cell(row=r, column=1, value="s.a:")
        r += 1
        snap0 = record["iteraciones"][0]
        for i, (fila, b) in enumerate(zip(snap0["A"], snap0["b"]), 1):
            izq = _str_combinacion(fila, record["nombres_var"])
            ws.cell(row=r, column=1, value=f"  {i}) {izq} = {fmt_num(b)}")
            r += 1

        # Ajustar anchos
        ws.column_dimensions['A'].width = 90
        for col in range(2, ncols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

    # ----------------------------------------------------------------
    def _split_coef_excel(self, coef):
        """Divide un Coef en (parte_constante, parte_M) para mostrar en dos celdas.

        Ej:  Coef(22000, +1)  -> ('22000', '+M')
             Coef(0, -1)      -> ('-1',    'M')
             Coef(-600, 0)    -> ('-600',  '')
             Coef(0, 0)       -> ('0',     '')
        """
        if not isinstance(coef, Coef):
            return (fmt_num(coef), "")
        if coef.m == 0:
            return (Coef._fmt(coef.const), "")
        # tiene componente M
        if coef.const == 0:
            # M puro: coeficiente en main, 'M' en sep
            return (Coef._fmt(coef.m), "M")
        # ambas componentes
        main = Coef._fmt(coef.const)
        if coef.m == 1:     sep = "+M"
        elif coef.m == -1:  sep = "-M"
        elif coef.m > 0:    sep = f"+{Coef._fmt(coef.m)}M"
        else:               sep = f"-{Coef._fmt(-coef.m)}M"
        return (main, sep)

    # ----------------------------------------------------------------
    def _add_iteracion(self, record, idx, snap):
        n_var  = record["n_total"]
        n_rest = len(snap["A"])
        nombres = record["nombres_var"]
        C = record["C"]

        nombre_hoja = f"Iteracion {idx}"
        if snap["es_optima"]:
            nombre_hoja += " (OPTIMA)"
        nombre_hoja = nombre_hoja[:31]
        ws = self.wb.create_sheet(nombre_hoja)

        # ---- Detectar qué columnas necesitan separador de M ----
        # Solo splitear si hay un valor MIXTO (constante != 0 Y m != 0),
        # como '22000+M'. Para M puros (-M, M), se muestra solo el coef en
        # la celda principal (formato del profesor).
        need_m_split = set()
        for j in range(n_var):
            zj = snap["Zj"][j]
            cz = snap["Cj_Zj"][j]
            zj_mixed = isinstance(zj, Coef) and zj.m != 0 and zj.const != 0
            cz_mixed = isinstance(cz, Coef) and cz.m != 0 and cz.const != 0
            if zj_mixed or cz_mixed:
                need_m_split.add(j)

        # ---- Mapeo de columnas Excel ----
        col_var0 = 4
        main_col = {}    # j -> columna principal
        sep_col  = {}    # j -> columna separador M (solo para j en need_m_split)
        current = col_var0
        for j in range(n_var):
            main_col[j] = current
            current += 1
            if j in need_m_split:
                sep_col[j] = current
                current += 1
        col_b   = current
        col_coc = col_b + 1
        ncols   = col_coc

        # Título
        titulo = f"ITERACION {idx}"
        if snap["es_optima"]:  titulo += "  --  TABLA OPTIMA"
        self._title_row(ws, 1, titulo, ncols=ncols)

        # Operaciones de fila previas
        r = 3
        if snap["ops_prev"]:
            ws.cell(row=r, column=1, value="Operaciones de fila aplicadas:").font = self.font_bold
            r += 1
            for op in snap["ops_prev"]:
                ws.cell(row=r, column=1, value=f"   {op}")
                r += 1
            r += 1

        # ---- Encabezado ----
        rh1 = r          # fila Cj
        rh2 = r + 1      # fila variables (X1, X2, ...)

        # Celda diagonal "Var. Cj / Básicas"
        ws.cell(row=rh1, column=1, value="Var. Cj")
        ws.cell(row=rh2, column=1, value="Básicas")
        for cell in (ws.cell(row=rh1, column=1), ws.cell(row=rh2, column=1)):
            cell.font = self.font_bold
            cell.alignment = self.center
            cell.fill = self.F_HEADER
        ws.cell(row=rh1, column=1).border = Border(
            left=self.THIN_S, right=self.THIN_S, top=self.THIN_S, bottom=None,
            diagonal=Side(style='thin'), diagonalDown=True
        )

        # Cb, V.B. headers
        ws.cell(row=rh1, column=2, value="").fill = self.F_HEADER
        ws.cell(row=rh1, column=3, value="").fill = self.F_HEADER
        ws.cell(row=rh2, column=2, value="Cb").font = self.font_bold
        ws.cell(row=rh2, column=3, value="V.B.").font = self.font_bold
        ws.cell(row=rh2, column=2).fill = self.F_HEADER
        ws.cell(row=rh2, column=3).fill = self.F_HEADER
        ws.cell(row=rh2, column=2).alignment = self.center
        ws.cell(row=rh2, column=3).alignment = self.center

        # Cj values y nombres de variables
        for j in range(n_var):
            c_main = main_col[j]
            ws.cell(row=rh1, column=c_main, value=str(C[j]))
            ws.cell(row=rh2, column=c_main, value=nombres[j])
            for rh in (rh1, rh2):
                cell = ws.cell(row=rh, column=c_main)
                cell.font = self.font_bold
                cell.alignment = self.center
                cell.fill = self.F_HEADER
            if snap["col_piv"] == j and not snap["es_optima"]:
                ws.cell(row=rh1, column=c_main).fill = self.F_ORANGE
                ws.cell(row=rh2, column=c_main).fill = self.F_ORANGE
            # Separador M (blanco en cabecera)
            if j in sep_col:
                c_sep = sep_col[j]
                for rh in (rh1, rh2):
                    ws.cell(row=rh, column=c_sep, value="").fill = self.F_HEADER
                if snap["col_piv"] == j and not snap["es_optima"]:
                    ws.cell(row=rh1, column=c_sep).fill = self.F_ORANGE
                    ws.cell(row=rh2, column=c_sep).fill = self.F_ORANGE

        # b y b/aij
        ws.cell(row=rh1, column=col_b, value="").fill = self.F_HEADER
        ws.cell(row=rh2, column=col_b, value="b").font = self.font_bold
        ws.cell(row=rh2, column=col_b).fill = self.F_HEADER
        ws.cell(row=rh2, column=col_b).alignment = self.center
        if not snap["es_optima"]:
            ws.cell(row=rh1, column=col_coc, value="").fill = self.F_HEADER
            ws.cell(row=rh2, column=col_coc, value="b/aij").font = self.font_bold
            ws.cell(row=rh2, column=col_coc).fill = self.F_HEADER
            ws.cell(row=rh2, column=col_coc).alignment = self.center

        # ---- Filas de datos ----
        r_data0 = rh2 + 1
        for i in range(n_rest):
            r_i = r_data0 + i
            ws.cell(row=r_i, column=1, value=f"F{i+1}").font = self.font_bold
            ws.cell(row=r_i, column=1).fill = self.F_HEADER
            ws.cell(row=r_i, column=1).alignment = self.center

            # Cb y V.B. (con split si la básica es artificial: Cb -> "-1", V.B. -> "MA5")
            cb_raw = snap["Cb"][i]
            vb_raw = snap["VB"][i]
            if isinstance(cb_raw, Coef) and cb_raw.m != 0 and cb_raw.const == 0:
                cb_display = Coef._fmt(cb_raw.m)   # -1 (Max) o +1 (Min)
                vb_display = "M" + vb_raw                # "MA5"
            else:
                cb_display = str(cb_raw)
                vb_display = vb_raw
            ws.cell(row=r_i, column=2, value=cb_display)
            ws.cell(row=r_i, column=3, value=vb_display)
            ws.cell(row=r_i, column=2).alignment = self.center
            ws.cell(row=r_i, column=3).alignment = self.center
            ws.cell(row=r_i, column=3).font = self.font_bold

            for j in range(n_var):
                c_main = main_col[j]
                val = fmt_num(snap["A"][i][j])
                ws.cell(row=r_i, column=c_main, value=val)
                ws.cell(row=r_i, column=c_main).alignment = self.center
                if snap["col_piv"] == j and not snap["es_optima"]:
                    ws.cell(row=r_i, column=c_main).fill = self.F_LIGHTORG
                if (snap["fila_piv"] == i and snap["col_piv"] == j
                        and not snap["es_optima"]):
                    ws.cell(row=r_i, column=c_main).fill = self.F_RED
                    ws.cell(row=r_i, column=c_main).font = self.font_white_bold
                # Separador M (vacío en filas de datos)
                if j in sep_col:
                    c_sep = sep_col[j]
                    ws.cell(row=r_i, column=c_sep, value="")
                    if snap["col_piv"] == j and not snap["es_optima"]:
                        ws.cell(row=r_i, column=c_sep).fill = self.F_LIGHTORG

            # b
            ws.cell(row=r_i, column=col_b, value=fmt_num(snap["b"][i]))
            ws.cell(row=r_i, column=col_b).alignment = self.center
            if snap["fila_piv"] == i and not snap["es_optima"]:
                ws.cell(row=r_i, column=col_b).font = self.font_bold
            # b/aij
            if snap["coc"]:
                ws.cell(row=r_i, column=col_coc, value=snap["coc"][i])
                ws.cell(row=r_i, column=col_coc).alignment = self.center
                if snap["fila_piv"] == i:
                    ws.cell(row=r_i, column=col_coc).font = self.font_bold
                    ws.cell(row=r_i, column=col_coc).fill = self.F_LIGHTORG

        # ---- Filas Zj y Cj-Zj (con split de M) ----
        r_zj   = r_data0 + n_rest
        r_diff = r_zj + 1
        ws.cell(row=r_zj,   column=3, value="Zj").font = self.font_bold
        ws.cell(row=r_diff, column=3, value="Cj-Zj").font = self.font_bold
        ws.cell(row=r_zj,   column=3).fill = self.F_HEADER
        ws.cell(row=r_diff, column=3).fill = self.F_HEADER
        ws.cell(row=r_zj,   column=3).alignment = self.center
        ws.cell(row=r_diff, column=3).alignment = self.center

        for j in range(n_var):
            c_main = main_col[j]
            zj_main, zj_sep = self._split_coef_excel(snap["Zj"][j])
            cz_main, cz_sep = self._split_coef_excel(snap["Cj_Zj"][j])
            ws.cell(row=r_zj,   column=c_main, value=zj_main)
            ws.cell(row=r_diff, column=c_main, value=cz_main)
            ws.cell(row=r_zj,   column=c_main).alignment = self.center
            ws.cell(row=r_diff, column=c_main).alignment = self.center
            if snap["col_piv"] == j and not snap["es_optima"]:
                ws.cell(row=r_diff, column=c_main).fill = self.F_ORANGE
                ws.cell(row=r_diff, column=c_main).font = self.font_bold
            # Separador M (lleva la parte M de Zj y Cj-Zj)
            if j in sep_col:
                c_sep = sep_col[j]
                ws.cell(row=r_zj,   column=c_sep, value=zj_sep)
                ws.cell(row=r_diff, column=c_sep, value=cz_sep)
                ws.cell(row=r_zj,   column=c_sep).alignment = self.center
                ws.cell(row=r_diff, column=c_sep).alignment = self.center
                if snap["col_piv"] == j and not snap["es_optima"]:
                    ws.cell(row=r_diff, column=c_sep).fill = self.F_ORANGE
                    ws.cell(row=r_diff, column=c_sep).font = self.font_bold

        # b (Z)
        ws.cell(row=r_zj, column=col_b, value=str(snap["Z"]))
        ws.cell(row=r_zj, column=col_b).font = self.font_bold
        ws.cell(row=r_zj, column=col_b).alignment = self.center

        # ---- Bordes para toda la tabla ----
        for rr in range(rh1, r_diff + 1):
            for cc in range(1, ncols + 1):
                cell = ws.cell(row=rr, column=cc)
                if not (rr == rh1 and cc == 1):
                    cell.border = self.border_all
        # Borde grueso top y bottom
        for cc in range(1, ncols + 1):
            ws.cell(row=rh1, column=cc).border = Border(
                left=ws.cell(row=rh1, column=cc).border.left,
                right=ws.cell(row=rh1, column=cc).border.right,
                top=self.MEDIUM_S,
                bottom=ws.cell(row=rh1, column=cc).border.bottom,
                diagonal=ws.cell(row=rh1, column=cc).border.diagonal,
                diagonalDown=ws.cell(row=rh1, column=cc).border.diagonalDown,
            )
            ws.cell(row=r_diff, column=cc).border = Border(
                left=ws.cell(row=r_diff, column=cc).border.left,
                right=ws.cell(row=r_diff, column=cc).border.right,
                top=ws.cell(row=r_diff, column=cc).border.top,
                bottom=self.MEDIUM_S,
            )

        # ---- Indicadores ----
        r_info = r_diff + 2
        if snap["col_piv"] is not None and snap["fila_piv"] is not None:
            ent = nombres[snap["col_piv"]]
            sal = snap["VB"][snap["fila_piv"]]
            piv = snap["A"][snap["fila_piv"]][snap["col_piv"]]
            dz  = snap["Cj_Zj"][snap["col_piv"]]
            tag = "MAYOR positivo" if record["tipo"] == "max" else "MAYOR negativo"
            ws.cell(row=r_info, column=1,
                    value=f"Variable que ENTRA:  {ent}  (Cj-Zj = {dz}, el {tag})")
            ws.cell(row=r_info, column=1).font = self.font_bold
            ws.cell(row=r_info + 1, column=1,
                    value=f"Variable que SALE :  {sal}  (menor cociente positivo b/aij)")
            ws.cell(row=r_info + 1, column=1).font = self.font_bold
            ws.cell(row=r_info + 2, column=1,
                    value=f"Elemento PIVOTE   :  a[F{snap['fila_piv']+1}][{ent}] = {fmt_num(piv)}")
            ws.cell(row=r_info + 2, column=1).font = self.font_bold
        elif snap["es_optima"]:
            ws.cell(row=r_info, column=1,
                    value=">>> Todos los Cj-Zj son <= 0  =>  TABLA OPTIMA <<<")
            ws.cell(row=r_info, column=1).font = Font(bold=True, color='006100')
            ws.cell(row=r_info, column=1).fill = self.F_GREEN

        # ---- Anchos de columnas ----
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 8     # Cb
        ws.column_dimensions['C'].width = 10    # V.B.
        for j in range(n_var):
            ws.column_dimensions[get_column_letter(main_col[j])].width = 12
            if j in sep_col:
                ws.column_dimensions[get_column_letter(sep_col[j])].width = 7
        ws.column_dimensions[get_column_letter(col_b)].width = 16
        if not snap["es_optima"]:
            ws.column_dimensions[get_column_letter(col_coc)].width = 12

    # ----------------------------------------------------------------
    def _add_solucion(self, record, metadatos=None):
        ws    = self.wb.create_sheet("Solucion")
        ncols = 7
        self._title_row(ws, 1, "SOLUCION DEL PROBLEMA", ncols=ncols)

        desc_vars = {}
        fo_desc   = ""
        if metadatos:
            desc_vars = metadatos.get("descripcion_variables", {})
            fo_desc   = metadatos.get("fo_descripcion", "")

        r = 3
        if record["estado"] == "optima":
            ws.cell(row=r, column=1, value="ESTADO: SOLUCION OPTIMA ENCONTRADA").font = \
                Font(bold=True, color='006100', size=12)
            ws.cell(row=r, column=1).fill = self.F_GREEN
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            r += 2

            # ── [1] Tabla de variables ──────────────────────────────────────
            hdrs = ["Variable", "Valor exacto", "Valor aprox.", "Tipo", "Descripcion / Interpretacion"]
            for ci, h in enumerate(hdrs, 1):
                cell = ws.cell(row=r, column=ci, value=h)
                cell.font      = self.font_bold
                cell.fill      = self.F_HEADER
                cell.alignment = self.center
                cell.border    = self.border_all
            r += 1

            valores     = record["valores"]
            nombres_dec = record["nombres_dec"]

            for n in nombres_dec:
                v    = valores[n]
                desc = desc_vars.get(n, "Variable de decision")
                ws.cell(row=r, column=1, value=n)
                ws.cell(row=r, column=2, value=fmt_num(v))
                ws.cell(row=r, column=3, value=round(float(v), 6))
                ws.cell(row=r, column=4, value="Decision")
                ws.cell(row=r, column=5, value=desc)
                ws.cell(row=r, column=5).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                for ci in range(1, 6):
                    ws.cell(row=r, column=ci).border = self.border_all
                    if ci != 5:
                        ws.cell(row=r, column=ci).alignment = self.center
                r += 1

            for j in range(record["n_dec"], record["n_total"]):
                n  = record["nombres_var"][j]
                tp = record["tipo_var"][j]
                v  = valores[n]
                num_rest = re.sub(r'[^0-9]', '', n)
                r_label  = f"Restriccion R{num_rest}" if num_rest else "Auxiliar"
                if tp == "A":
                    if v != 0:
                        tipo_txt = "Artificial (revisar)"
                        coment   = "Variable artificial > 0: verificar factibilidad"
                    else:
                        continue
                elif tp == "S+":
                    tipo_txt = "Holgura"
                    coment   = (f"Recurso sobrante en {r_label} ({fmt_num(v)} unidades libres)"
                                if v != 0 else f"{r_label} ACTIVA — recurso totalmente agotado")
                else:
                    tipo_txt = "Excedente"
                    coment   = (f"Se supera el minimo de {r_label} en {fmt_num(v)} unidades"
                                if v != 0 else f"{r_label} ACTIVA — minimo cumplido exactamente")
                ws.cell(row=r, column=1, value=n)
                ws.cell(row=r, column=2, value=fmt_num(v))
                ws.cell(row=r, column=3, value=round(float(v), 6))
                ws.cell(row=r, column=4, value=tipo_txt)
                ws.cell(row=r, column=5, value=coment)
                ws.cell(row=r, column=5).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                for ci in range(1, 6):
                    ws.cell(row=r, column=ci).border = self.border_all
                    if ci != 5:
                        ws.cell(row=r, column=ci).alignment = self.center
                r += 1

            # Fila Z
            r += 1
            Z        = record["Z_final"]
            verbo_z  = "Valor MAXIMO de la F.O." if record["tipo"] == "max" else "Valor MINIMO de la F.O."
            if fo_desc:
                verbo_z += f"  ({fo_desc})"
            ws.cell(row=r, column=1, value="Z").font = Font(bold=True, size=12)
            ws.cell(row=r, column=2, value=str(Z)).font = Font(bold=True, size=12)
            try:
                zval = float(Z.const) if isinstance(Z, Coef) and Z.m == 0 else float(Z)
                ws.cell(row=r, column=3, value=round(zval, 6)).font = Font(bold=True, size=12)
            except Exception:
                pass
            ws.cell(row=r, column=4, value="F.O.").font = Font(bold=True, size=12)
            ws.cell(row=r, column=5, value=verbo_z).font = Font(bold=True, size=12)
            ws.cell(row=r, column=5).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            for ci in range(1, 6):
                ws.cell(row=r, column=ci).fill      = self.F_GREEN
                ws.cell(row=r, column=ci).border    = self.border_med
                if ci != 5:
                    ws.cell(row=r, column=ci).alignment = self.center
            r += 3

            # ── [2] Verificacion de restricciones ──────────────────────────
            ws.cell(row=r, column=1, value="VERIFICACION DE RESTRICCIONES (valores optimos sustituidos)")
            ws.cell(row=r, column=1).font = Font(bold=True, size=11)
            ws.cell(row=r, column=1).fill = self.F_HEADER
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            r += 1
            for ci, h in enumerate(["Rest.", "Calculo (LHS)", "Resultado LHS", "Signo", "RHS", "Estado"], 1):
                cell = ws.cell(row=r, column=ci, value=h)
                cell.font      = self.font_bold
                cell.fill      = self.F_HEADER
                cell.alignment = self.center
                cell.border    = self.border_all
            r += 1
            for i, (a, s, b) in enumerate(record["restricciones_orig"], 1):
                lhs_val = sum(Fraction(a[j]) * valores[nombres_dec[j]] for j in range(len(nombres_dec)))
                terms   = [f"({fmt_num(Fraction(a[j]))})\u00d7({fmt_num(valores[nombres_dec[j]])})"
                           for j in range(len(nombres_dec)) if Fraction(a[j]) != 0]
                lhs_str = " + ".join(terms) if terms else "0"
                ok      = ((s == "<=" and lhs_val <= b) or
                           (s == ">=" and lhs_val >= b) or
                           (s == "="  and lhs_val == b))
                est_txt  = "SATISFECHA" if ok else "NO SATISFECHA"
                est_fill = self.F_GREEN if ok else PatternFill('solid', fgColor='FFC7CE')
                ws.cell(row=r, column=1, value=f"R{i}")
                ws.cell(row=r, column=2, value=lhs_str)
                ws.cell(row=r, column=3, value=fmt_num(lhs_val))
                ws.cell(row=r, column=4, value=s)
                ws.cell(row=r, column=5, value=fmt_num(b))
                ws.cell(row=r, column=6, value=est_txt)
                ws.cell(row=r, column=6).fill = est_fill
                ws.cell(row=r, column=6).font = self.font_bold
                for ci in range(1, 7):
                    ws.cell(row=r, column=ci).border    = self.border_all
                    ws.cell(row=r, column=ci).alignment = self.center
                ws.cell(row=r, column=2).alignment = Alignment(horizontal='left', wrap_text=True)
                r += 1
            r += 2

            # ── [3] Conclusion ─────────────────────────────────────────────
            ws.cell(row=r, column=1, value="CONCLUSION / INTERPRETACION")
            ws.cell(row=r, column=1).font = Font(bold=True, size=11)
            ws.cell(row=r, column=1).fill = self.F_HEADER
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            r += 1
            accion = "MAXIMIZAR" if record["tipo"] == "max" else "MINIMIZAR"
            lines_concl = [f"Para {accion} la funcion objetivo Z, la solucion optima indica:"]
            for n in nombres_dec:
                v    = valores[n]
                desc = desc_vars.get(n, "")
                ap2  = f" (~{float(v):.4f})" if (isinstance(v, Fraction) and v.denominator != 1) else ""
                tag  = f"  ->  {desc}" if desc else ""
                lines_concl.append(f"  * {n} = {fmt_num(v)}{ap2}{tag}")
            lines_concl.append("")
            activas     = []
            con_holgura = []
            for j in range(record["n_dec"], record["n_total"]):
                nv  = record["nombres_var"][j]
                tp  = record["tipo_var"][j]
                vv  = valores[nv]
                if tp not in ("S+", "S-"):
                    continue
                nm = re.sub(r'[^0-9]', '', nv)
                if not nm:
                    continue
                idx_r = int(nm)
                if vv == 0:
                    activas.append(idx_r)
                else:
                    tipo_h = "holgura" if tp == "S+" else "excedente"
                    con_holgura.append((idx_r, fmt_num(vv), tipo_h))
            if activas:
                r_str = ", ".join(f"R{r_}" for r_ in sorted(activas))
                lines_concl.append(f"Las restricciones {r_str} son ACTIVAS (se usan al 100% de su capacidad).")
            for idx_r, val, tipo_h in sorted(con_holgura):
                lines_concl.append(f"La restriccion R{idx_r} tiene {tipo_h} de {val} unidades.")
            verbo_obj2 = "maximo" if record["tipo"] == "max" else "minimo"
            lines_concl.append(f"\nEl valor {verbo_obj2} de la funcion objetivo es:  Z = {Z}")
            concl_text = "\n".join(lines_concl)
            n_lines    = len(lines_concl) + 2
            ws.cell(row=r, column=1, value=concl_text)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r + n_lines, end_column=ncols)
            ws.row_dimensions[r].height = 15 * n_lines

        elif record["estado"] == "infactible":
            ws.cell(row=r, column=1,
                    value="ESTADO: PROBLEMA INFACTIBLE").font = Font(bold=True, color='9C0006', size=12)
            ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor='FFC7CE')
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            r += 1
            ws.cell(row=r, column=1,
                    value="Una variable artificial quedo en la base con valor > 0.")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)

        elif record["estado"] == "no_acotada":
            ws.cell(row=r, column=1,
                    value="ESTADO: PROBLEMA NO ACOTADO").font = Font(bold=True, color='9C5700', size=12)
            ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor='FFEB9C')
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            r += 1
            ws.cell(row=r, column=1,
                    value="La funcion objetivo puede crecer/decrecer sin limite.")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)

        # Anchos de columnas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 52
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 14


# ============================================================================
#  INTERFAZ GRÁFICA (tkinter)
# ============================================================================
EJEMPLO_DEFAULT = """\
Max  Z = 22000 X1 + 20000 X2
s.a.
   2,5 X1 + 3 X2  <=  4500
   3 X1 + 6 X2    <=  8400
   14 X1 + 10 X2  <=  20000
   X1 + X2        <=  1700
   X1             >=  600
   X1, X2 >= 0
"""

EJEMPLO_MIN = """\
Min  Z = 3 X1 + 4 X2
s.a.
   4 X1 + 4 X2   <=  16
   4 X1 + 12 X2  >=  24
   X1, X2 >= 0
"""

EJEMPLO_PINTURA = """\
Max  Z = 3000 X1 + 2000 X2
s.a.
   X1 + 2 X2   <=  6
   2 X1 + X2   <=  8
   -X1 + X2    <=  1
   X2          <=  2
   X1, X2 >= 0
"""


class SimplexApp:

    def __init__(self, root):
        self.root = root
        self.root.title("Resolutor del Método Simplex (Gran M) - Brayan")
        self.root.geometry("1180x780")
        self.root.minsize(900, 600)

        self.last_record = None
        self.last_enunciado = ""
        self.last_metadatos = None
        self._embedded_widgets = []
        self.config = cargar_config()
        self.api_key = self.config.get("groq_api_key", "")
        self.api_model = self.config.get("groq_model", "llama-3.3-70b-versatile")
        self._build_ui()

    def _build_ui(self):
        # Título
        title_frame = tk.Frame(self.root, bg='#305496', height=50)
        title_frame.pack(fill='x', side='top')
        title_frame.pack_propagate(False)
        tk.Label(title_frame,
                 text="  MÉTODO SIMPLEX - Gran M  ",
                 font=('Segoe UI', 16, 'bold'),
                 bg='#305496', fg='white').pack(side='left', padx=10, pady=10)
        tk.Label(title_frame,
                 text="Investigación de Operaciones",
                 font=('Segoe UI', 11),
                 bg='#305496', fg='white').pack(side='left', padx=10)

        # Frame principal dividido
        main = tk.PanedWindow(self.root, orient='horizontal', sashwidth=6, bg='#d0d0d0')
        main.pack(fill='both', expand=True, padx=6, pady=6)

        # ----- Panel izquierdo: input -----
        left = tk.Frame(main, bg='#f0f0f0')
        main.add(left, minsize=350)

        tk.Label(left, text="ENUNCIADO o MODELO MATEMÁTICO (pega cualquiera aquí):",
                 font=('Segoe UI', 10, 'bold'), bg='#f0f0f0',
                 anchor='w').pack(fill='x', padx=4, pady=(6, 2))
        self.txt_input = scrolledtext.ScrolledText(left, font=('Consolas', 11),
                                                    wrap='word', undo=True,
                                                    bg='#1e1e2e', fg='#cdd6f4',
                                                    insertbackground='#cdd6f4')
        self.txt_input.pack(fill='both', expand=True, padx=4)
        self.txt_input.insert('1.0', EJEMPLO_DEFAULT)

        # Campo opcional de descripción de variables
        desc_frame = tk.Frame(left, bg='#f0f0f0')
        desc_frame.pack(fill='x', padx=4, pady=(3, 0))
        tk.Label(desc_frame,
                 text="Descrip. de variables (ej: X1 = Cámaras tipo A; X2 = Cámaras tipo B):",
                 font=('Segoe UI', 8), bg='#f0f0f0', fg='#666').pack(anchor='w')
        self.txt_desc_vars = ttk.Entry(desc_frame, font=('Consolas', 9))
        self.txt_desc_vars.pack(fill='x', expand=True)

        # Botones - Fila 1: acciones principales
        btn_frame = tk.Frame(left, bg='#f0f0f0')
        btn_frame.pack(fill='x', padx=4, pady=6)

        ttk.Style().configure('Big.TButton', font=('Segoe UI', 10, 'bold'))
        ttk.Button(btn_frame, text="▶ RESOLVER", style='Big.TButton',
                   command=self.on_resolver).pack(side='left', padx=2)
        ttk.Button(btn_frame, text="🤖 Interpretar con IA",
                   command=self.on_interpretar_ia).pack(side='left', padx=2)
        ttk.Button(btn_frame, text="📋 Cargar ejemplo ▾",
                   command=self.on_ejemplo).pack(side='left', padx=2)
        ttk.Button(btn_frame, text="📥 Cargar archivo",
                   command=self.on_cargar_archivo).pack(side='left', padx=2)
        ttk.Button(btn_frame, text="🧹 Limpiar",
                   command=self.on_limpiar).pack(side='left', padx=2)

        # Botones - Fila 2: exportar y config
        btn_frame2 = tk.Frame(left, bg='#f0f0f0')
        btn_frame2.pack(fill='x', padx=4, pady=(0, 6))
        self.btn_exportar = ttk.Button(btn_frame2, text="📊 Exportar a Excel",
                                        command=self.on_exportar, state='disabled')
        self.btn_exportar.pack(side='left', padx=2)
        self.btn_guardar_txt = ttk.Button(btn_frame2, text="💾 Guardar resultado (.txt)",
                                           command=self.on_guardar_txt, state='disabled')
        self.btn_guardar_txt.pack(side='left', padx=2)
        ttk.Button(btn_frame2, text="🔑 Configurar API key (IA)",
                   command=self.on_config_api).pack(side='right', padx=2)

        # Hint del formato
        hint = tk.Label(left, font=('Consolas', 8), bg='#f0f0f0', fg='#555',
                        anchor='w', justify='left',
                        text="RESOLVER detecta automaticamente si es enunciado natural o modelo matematico.\n"
                             "Modelo:  Max/Min Z = ... ;  <= , >= , = ;  X1, X2, ... ;  coma o punto decimal; fracciones 5/2 OK\n"
                             "Enunciado natural: requiere API key de Groq (boton 'Configurar API key').")
        hint.pack(fill='x', padx=4, pady=(0, 4))

        # ----- Panel derecho: output -----
        right = tk.Frame(main, bg='#f0f0f0')
        main.add(right, minsize=500)

        tk.Label(right, text="RESULTADO Y PROCEDIMIENTO PASO A PASO:",
                 font=('Segoe UI', 10, 'bold'), bg='#f0f0f0',
                 anchor='w').pack(fill='x', padx=4, pady=(6, 2))
        self.txt_output = scrolledtext.ScrolledText(right, font=('Consolas', 10),
                                                     wrap='none', state='disabled',
                                                     bg='#1e1e2e', fg='#cdd6f4',
                                                     insertbackground='#cdd6f4')
        self.txt_output.pack(fill='both', expand=True, padx=4, pady=(0, 4))

        # Status bar
        self.status = tk.Label(self.root, text="Listo. Pulsa RESOLVER.",
                                bd=1, relief='sunken', anchor='w',
                                bg='#e0e0e0', font=('Segoe UI', 9))
        self.status.pack(side='bottom', fill='x')

        if not OPENPYXL_OK:
            self.status.config(text="AVISO: openpyxl no instalado -> 'Exportar a Excel' deshabilitado. Instala con: pip install openpyxl",
                                fg='#b00')

    # ----- Acciones -----
    def _clear_output(self):
        """Limpia el área de salida y destruye los widgets de tabla embebidos."""
        for w in self._embedded_widgets:
            try:
                w.destroy()
            except Exception:
                pass
        self._embedded_widgets = []
        self.txt_output.config(state='normal')
        self.txt_output.delete('1.0', 'end')
        self.txt_output.config(state='disabled')

    def _set_output(self, texto):
        self._clear_output()
        self.txt_output.config(state='normal')
        self.txt_output.insert('1.0', texto)
        self.txt_output.config(state='disabled')

    def _build_table_widget(self, record, snap):
        """Construye un Frame con la tabla simplex como tabla real (celdas coloreadas)."""
        n_var    = record["n_total"]
        n_rest   = len(snap["A"])
        nombres  = record["nombres_var"]
        C        = record["C"]
        col_piv  = snap.get("col_piv")
        fila_piv = snap.get("fila_piv")
        es_opt   = snap["es_optima"]

        # Paleta de colores (tema oscuro)
        CH_BG = '#363659'; CH_FG = '#cdd6f4'   # cabecera
        CD_BG = '#252540'; CD_FG = '#cdd6f4'   # datos
        CA_BG = '#2d2d52'                       # datos alternado
        CP_BG = '#4a3000'; CP_FG = '#ffc060'   # columna pivote
        CE_BG = '#8b0000'; CE_FG = '#ffffff'   # celda pivote (elemento)
        DARK  = '#1e1e2e'

        outer = tk.Frame(self.txt_output, bg=DARK, padx=4, pady=4)
        # bg del frame interior actúa como color de borde entre celdas
        f = tk.Frame(outer, bg='#55558a')
        f.pack()

        def cell(row, col, text, bg, fg, bold=False):
            lbl = tk.Label(f, text=str(text), bg=bg, fg=fg,
                           font=('Consolas', 9, 'bold' if bold else 'normal'),
                           padx=7, pady=3, anchor='center')
            lbl.grid(row=row, column=col, sticky='nsew', padx=1, pady=1)

        # ── Fila 0: valores Cj ──
        cell(0, 0, "",     CH_BG, CH_FG)
        cell(0, 1, "",     CH_BG, CH_FG)
        cell(0, 2, "Cj→", CH_BG, CH_FG, bold=True)
        for j in range(n_var):
            is_pc = not es_opt and col_piv == j
            cell(0, j+3, str(C[j]),
                 CP_BG if is_pc else CH_BG,
                 CP_FG if is_pc else CH_FG, bold=True)
        cell(0, n_var+3, "",      CH_BG, CH_FG)
        cell(0, n_var+4, "",      CH_BG, CH_FG)

        # ── Fila 1: nombres de columnas ──
        cell(1, 0, "Fila",  CH_BG, CH_FG, bold=True)
        cell(1, 1, "Cb",    CH_BG, CH_FG, bold=True)
        cell(1, 2, "V.B.",  CH_BG, CH_FG, bold=True)
        for j in range(n_var):
            is_pc = not es_opt and col_piv == j
            cell(1, j+3, nombres[j],
                 CP_BG if is_pc else CH_BG,
                 CP_FG if is_pc else CH_FG, bold=True)
        cell(1, n_var+3, "b",     CH_BG, CH_FG, bold=True)
        cell(1, n_var+4, "b/aij", CH_BG, CH_FG, bold=True)

        # ── Filas de datos ──
        for i in range(n_rest):
            is_pr   = not es_opt and fila_piv == i
            row_bg  = CA_BG if i % 2 == 0 else CD_BG
            fila_lbl = f"►F{i+1}" if is_pr else f"F{i+1}"
            cell(i+2, 0, fila_lbl,
                 CP_BG if is_pr else CH_BG,
                 CP_FG if is_pr else CH_FG, bold=is_pr)
            cell(i+2, 1, str(snap["Cb"][i]), row_bg, CD_FG)
            cell(i+2, 2, snap["VB"][i],      row_bg, CD_FG, bold=True)
            for j in range(n_var):
                is_pc = not es_opt and col_piv == j
                is_pe = is_pr and is_pc
                val   = fmt_num(snap["A"][i][j])
                if   is_pe: bg, fg, bd = CE_BG, CE_FG, True
                elif is_pc: bg, fg, bd = CP_BG, CP_FG, False
                else:       bg, fg, bd = row_bg, CD_FG, False
                cell(i+2, j+3, val, bg, fg, bd)
            cell(i+2, n_var+3, fmt_num(snap["b"][i]),
                 CP_BG if is_pr else row_bg,
                 CP_FG if is_pr else CD_FG, bold=is_pr)
            coc_val = snap["coc"][i] if snap["coc"] else ""
            cell(i+2, n_var+4, coc_val,
                 CP_BG if is_pr else row_bg,
                 CP_FG if is_pr else CD_FG, bold=is_pr)

        # ── Fila Zj ──
        zj_r = n_rest + 2
        cell(zj_r, 0, "",      CH_BG, CH_FG)
        cell(zj_r, 1, "",      CH_BG, CH_FG)
        cell(zj_r, 2, "Zj",   CH_BG, CH_FG, bold=True)
        for j in range(n_var):
            cell(zj_r, j+3, str(snap["Zj"][j]), CH_BG, CH_FG)
        cell(zj_r, n_var+3, str(snap["Z"]), CH_BG, CH_FG, bold=True)
        cell(zj_r, n_var+4, "",             CH_BG, CH_FG)

        # ── Fila Cj-Zj ──
        cz_r = n_rest + 3
        cell(cz_r, 0, "",       CH_BG, CH_FG)
        cell(cz_r, 1, "",       CH_BG, CH_FG)
        cell(cz_r, 2, "Cj-Zj", CH_BG, CH_FG, bold=True)
        for j in range(n_var):
            is_pc = not es_opt and col_piv == j
            cell(cz_r, j+3, str(snap["Cj_Zj"][j]),
                 '#8a5c00' if is_pc else CH_BG,
                 '#ffffff' if is_pc else CH_FG, bold=is_pc)
        cell(cz_r, n_var+3, "", CH_BG, CH_FG)
        cell(cz_r, n_var+4, "", CH_BG, CH_FG)

        # Igualar anchos de columna
        for c in range(n_var + 5):
            f.grid_columnconfigure(c, weight=1, uniform='tbl')

        return outer

    def _insert_table(self, record, snap):
        """Embebe un widget de tabla real dentro del ScrolledText."""
        tbl = self._build_table_widget(record, snap)
        self._embedded_widgets.append(tbl)
        self.txt_output.window_create('end', window=tbl, padx=4, pady=2)
        self.txt_output.insert('end', '\n')

    def _render_output(self, record, metadatos=None, preface=""):
        """Renderiza el procedimiento completo: texto + tablas como widgets reales."""
        self._clear_output()
        self.txt_output.config(state='normal')
        T = self.txt_output

        def txt(s):
            T.insert('end', s)

        ANCHO = 78
        SEP   = "=" * ANCHO
        verbo_max   = "MAXIMIZACION" if record["tipo"] == "max" else "MINIMIZACION"
        verbo_short = "Max" if record["tipo"] == "max" else "Min"

        if preface:
            txt(preface + "\n\n")

        # Encabezado
        txt(SEP + "\n")
        txt(f"  RESOLUCION POR EL METODO SIMPLEX (Gran M) - {verbo_max}\n")
        txt(SEP + "\n\n")

        # Modelo general
        txt("-" * ANCHO + "\n MODELO GENERAL\n" + "-" * ANCHO + "\n")
        fo = _str_combinacion(record["c_orig"], record["nombres_dec"])
        txt(f"  F.O :  {verbo_short}.  Z = {fo}\n  s.a (sujeto a):\n")
        for i, (a, s, b) in enumerate(record["restricciones_orig"], 1):
            txt(f"    {i}) {_str_combinacion(a, record['nombres_dec'])} {s} {fmt_num(b)}\n")
        txt(f"  {', '.join(record['nombres_dec'])} >= 0\n")

        # Modelo estándar
        txt("\n" + "-" * ANCHO + "\n MODELO ESTANDAR (forma canonica del Simplex)\n" + "-" * ANCHO + "\n")
        txt("  Reglas usadas:\n")
        txt("    - Cada '<='  recibe HOLGURA  S (coef +1).\n")
        txt("    - Cada '>='  recibe EXCEDENTE S (coef -1) + ARTIFICIAL A (coef +1).\n")
        txt("    - Cada '='   recibe ARTIFICIAL A.\n")
        txt("    - En la F.O., las A llevan -M (Max) o +M (Min).\n\n")
        fo_e = _str_combinacion(record["C"], record["nombres_var"], incluir_ceros=True)
        txt(f"  F.O :  {verbo_short}.  Z = {fo_e}\n  s.a:\n")
        snap0 = record["iteraciones"][0]
        for i, (fila, b) in enumerate(zip(snap0["A"], snap0["b"]), 1):
            txt(f"    {i}) {_str_combinacion(fila, record['nombres_var'])} = {fmt_num(b)}\n")

        # Criterios
        txt("\n" + "-" * ANCHO + "\n CRITERIOS DEL METODO\n" + "-" * ANCHO + "\n")
        if record["tipo"] == "max":
            txt("  - VARIABLE QUE ENTRA : la del MAYOR Cj-Zj POSITIVO.\n")
            txt("  - OPTIMO            : todos los Cj-Zj son <= 0.\n")
        else:
            txt("  - VARIABLE QUE ENTRA : la del MAYOR Cj-Zj NEGATIVO.\n")
            txt("  - OPTIMO            : todos los Cj-Zj son >= 0.\n")
        txt("  - VARIABLE QUE SALE  : la del MENOR cociente POSITIVO b/aij.\n")
        txt("  - PIVOTEO            : F_piv <- F_piv / piv ; F_otra <- F_otra - aij * F_piv.\n")

        # ── Iteraciones ──
        for it_idx, snap in enumerate(record["iteraciones"]):
            txt("\n")
            if snap["es_optima"]:
                txt("+" + "-" * (ANCHO - 2) + "+\n")
                txt(f"| ITERACION {it_idx}  --  TABLA OPTIMA" + " " * (ANCHO - 38) + "|\n")
                txt("+" + "-" * (ANCHO - 2) + "+\n")
            else:
                txt("+" + "-" * (ANCHO - 2) + "+\n")
                txt(f"| ITERACION {it_idx}" + " " * (ANCHO - 16) + "|\n")
                txt("+" + "-" * (ANCHO - 2) + "+\n")

            if snap["ops_prev"]:
                txt("  Operaciones de fila aplicadas:\n")
                for op in snap["ops_prev"]:
                    txt(f"    {op}\n")
                txt("\n")

            # ── TABLA COMO WIDGET REAL ──
            self._insert_table(record, snap)

            if not snap["es_optima"]:
                ent = record["nombres_var"][snap["col_piv"]]
                sal = snap["VB"][snap["fila_piv"]]
                piv = snap["A"][snap["fila_piv"]][snap["col_piv"]]
                dz  = snap["Cj_Zj"][snap["col_piv"]]
                tag = "MAYOR positivo" if record["tipo"] == "max" else "MAYOR negativo"
                txt(f"\n  >>> Variable que ENTRA: {ent}  (Cj-Zj = {dz}, el {tag})\n")
                txt(f"  >>> Variable que SALE : {sal}  (menor cociente positivo b/aij)\n")
                txt(f"  >>> Pivote            : a[F{snap['fila_piv']+1}][{ent}]"
                    f" = {fmt_num(piv)}  (marcado en rojo)\n")

        # ── Solución ──
        txt("\n" + SEP + "\n")
        if record["estado"] == "optima":
            valores     = record["valores"]
            nombres_dec = record["nombres_dec"]
            desc_vars   = {}
            fo_desc     = ""
            if metadatos:
                desc_vars = metadatos.get("descripcion_variables", {})
                fo_desc   = metadatos.get("fo_descripcion", "")
            SUB = "  " + "-" * (ANCHO - 4)

            txt("  SOLUCION OPTIMA ENCONTRADA\n" + SEP + "\n")

            if desc_vars:
                txt("\n  [1] IDENTIFICACION DE VARIABLES DE DECISION\n" + SUB + "\n")
                for n in nombres_dec:
                    desc = desc_vars.get(n, "")
                    txt(f"  {n:<6} =  {desc}\n" if desc else f"  {n}\n")

            txt("\n  [2] VALORES OPTIMOS DE LAS VARIABLES DE DECISION\n" + SUB + "\n")
            for n in nombres_dec:
                v   = valores[n]
                ap  = f"  (~{float(v):.4f})" if (isinstance(v, Fraction) and v.denominator != 1) else ""
                desc = desc_vars.get(n, "")
                tag  = f"   [{desc}]" if desc else ""
                txt(f"  {n} = {fmt_num(v)}{ap}{tag}\n")

            txt("\n  [3] ANALISIS DE HOLGURAS / EXCEDENTES\n" + SUB + "\n")
            alguna = False
            for j in range(record["n_dec"], record["n_total"]):
                n  = record["nombres_var"][j]
                tp = record["tipo_var"][j]
                v  = valores[n]
                if tp == "A":
                    continue
                alguna = True
                num_rest = re.sub(r'[^0-9]', '', n)
                r_label  = f"R{num_rest}" if num_rest else "?"
                ap = f"  (~{float(v):.4f})" if (isinstance(v, Fraction) and v.denominator != 1) else ""
                if tp == "S+":
                    if v == 0:
                        txt(f"  {n} = 0   -> Restriccion {r_label}: ACTIVA (recurso agotado, sin holgura)\n")
                    else:
                        txt(f"  {n} = {fmt_num(v)}{ap}   -> Restriccion {r_label}: holgura de {fmt_num(v)} unidades\n")
                elif tp == "S-":
                    if v == 0:
                        txt(f"  {n} = 0   -> Restriccion {r_label}: ACTIVA (se cumple exactamente el minimo exigido)\n")
                    else:
                        txt(f"  {n} = {fmt_num(v)}{ap}   -> Restriccion {r_label}: excedente de {fmt_num(v)} unidades sobre el minimo\n")
            if not alguna:
                txt("  (No hay variables de holgura/excedente en este modelo)\n")

            txt("\n  [4] VERIFICACION DE RESTRICCIONES (valores optimos sustituidos)\n" + SUB + "\n")
            for i, (a, s, b) in enumerate(record["restricciones_orig"], 1):
                lhs_val = sum(Fraction(a[j]) * valores[nombres_dec[j]] for j in range(len(nombres_dec)))
                terms   = [f"({fmt_num(Fraction(a[j]))})\u00d7({fmt_num(valores[nombres_dec[j]])})"
                           for j in range(len(nombres_dec)) if Fraction(a[j]) != 0]
                lhs_str = " + ".join(terms) if terms else "0"
                ok = ((s == "<=" and lhs_val <= b) or (s == ">=" and lhs_val >= b) or (s == "=" and lhs_val == b))
                txt(f"  R{i}: {lhs_str} = {fmt_num(lhs_val)}  {s}  {fmt_num(b)}   [{'SATISFECHA' if ok else 'NO SATISFECHA'}]\n")

            txt("\n  [5] VALOR DE LA FUNCION OBJETIVO\n" + SUB + "\n")
            Z  = record["Z_final"]
            ap = ""
            if isinstance(Z, Coef) and Z.m == 0 and Z.const.denominator != 1:
                ap = f"  (~{float(Z.const):.4f})"
            verbo_obj = "MAXIMO" if record["tipo"] == "max" else "MINIMO"
            fo_tag    = f"  ({fo_desc})" if fo_desc else ""
            txt(f"  Z = {Z}{ap}   <- Valor {verbo_obj} de la F.O.{fo_tag}\n")

            txt("\n  [6] CONCLUSION / INTERPRETACION\n" + SUB + "\n")
            accion = "MAXIMIZAR" if record["tipo"] == "max" else "MINIMIZAR"
            txt(f"  Para {accion} la funcion objetivo Z, la solucion optima indica:\n\n")
            for n in nombres_dec:
                v    = valores[n]
                desc = desc_vars.get(n, "")
                ap2  = f"  (~{float(v):.4f})" if (isinstance(v, Fraction) and v.denominator != 1) else ""
                tag  = f"  ->  {desc}" if desc else ""
                txt(f"    * {n} = {fmt_num(v)}{ap2}{tag}\n")
            txt("\n")
            activas     = []
            con_holgura = []
            for j in range(record["n_dec"], record["n_total"]):
                nv = record["nombres_var"][j]
                tp = record["tipo_var"][j]
                vv = valores[nv]
                if tp not in ("S+", "S-"):
                    continue
                nm = re.sub(r'[^0-9]', '', nv)
                if not nm:
                    continue
                idx_r = int(nm)
                if vv == 0:
                    activas.append(idx_r)
                else:
                    tipo_h = "holgura" if tp == "S+" else "excedente"
                    con_holgura.append((idx_r, fmt_num(vv), tipo_h))
            if activas:
                r_str = ", ".join(f"R{r_}" for r_ in sorted(activas))
                txt(f"  Las restricciones {r_str} son ACTIVAS (se usan al 100% de su capacidad).\n")
            for idx_r, val, tipo_h in sorted(con_holgura):
                txt(f"  La restriccion R{idx_r} tiene {tipo_h} de {val} unidades (capacidad no utilizada del todo).\n")
            verbo_obj2 = "maximo" if record["tipo"] == "max" else "minimo"
            txt(f"\n  El valor {verbo_obj2} de la funcion objetivo es  Z = {Z}.\n")
            txt(SEP + "\n")

        elif record["estado"] == "infactible":
            txt(" RESULTADO: PROBLEMA INFACTIBLE\n" + SEP + "\n")
            txt("  Una variable artificial quedó en la base con valor > 0.\n")
            txt("  Las restricciones no tienen solución factible.\n")
        elif record["estado"] == "no_acotada":
            txt(" RESULTADO: PROBLEMA NO ACOTADO\n" + SEP + "\n")
            txt("  La columna pivote no tiene ningún coeficiente positivo.\n")
            txt("  La función objetivo puede crecer/decrecer sin límite.\n")
        else:
            txt(" LIMITE DE ITERACIONES ALCANZADO\n" + SEP + "\n")

        T.config(state='disabled')
        T.see('1.0')

    def on_resolver(self):
        texto = self.txt_input.get('1.0', 'end').strip()
        if not texto:
            messagebox.showwarning("Sin modelo", "Por favor escribe el modelo o enunciado.")
            return

        # Auto-detectar: si parece enunciado natural, usar IA
        if es_enunciado_natural(texto):
            if not self.api_key:
                resp = messagebox.askyesno(
                    "Enunciado en lenguaje natural detectado",
                    "El texto parece ser un enunciado en lenguaje natural (no un modelo matemático).\n\n"
                    "Para interpretarlo automáticamente se requiere una API key de Groq.\n\n"
                    "¿Deseas configurar tu API key ahora?\n"
                    "(Alternativamente, puedes escribir el modelo matemático directamente.)")
                if resp:
                    self.on_config_api()
                if not self.api_key:
                    return
            self._resolver_con_ia(texto)
        else:
            self._resolver_modelo(texto)

    def _resolver_modelo(self, texto, metadatos=None):
        """Resuelve un modelo matemático (formato formal)."""
        try:
            self.status.config(text="Procesando modelo...", fg='#000')
            self.root.update_idletasks()
            tipo, c, restricciones, var_names = parse_problem(texto)
            self._ejecutar_solver(tipo, c, restricciones, var_names, texto, metadatos)
        except Exception as e:
            messagebox.showerror("Error al resolver",
                                  f"No se pudo procesar el modelo.\n\nDetalle:\n{e}")
            self.status.config(text=f"Error: {e}", fg='#b00')

    def _resolver_con_ia(self, texto):
        """Interpreta enunciado con IA y resuelve."""
        try:
            self.status.config(text=f"Consultando IA ({self.api_model}) para interpretar el enunciado...", fg='#000')
            self.root.update_idletasks()
            tipo, c, restricciones, var_names, metadatos = parse_enunciado_ia(
                texto, self.api_key, model=self.api_model)
            self._ejecutar_solver(tipo, c, restricciones, var_names, texto, metadatos)
        except Exception as e:
            messagebox.showerror("Error con la IA",
                                  f"No se pudo interpretar el enunciado.\n\nDetalle:\n{e}")
            self.status.config(text=f"Error IA: {e}", fg='#b00')

    def _leer_descripciones_vars(self):
        """Lee y parsea las descripciones del campo de texto de la UI.
        Formato soportado: 'X1 = Cámaras tipo A; X2 = Cámaras tipo B'
        Devuelve dict {nombre_var: descripcion}.
        """
        try:
            text = self.txt_desc_vars.get().strip()
        except Exception:
            return {}
        if not text:
            return {}
        desc = {}
        for parte in re.split(r'[;\n]', text):
            m = re.match(r'\s*([A-Za-z]\w*)\s*[=:]\s*(.+)', parte.strip())
            if m:
                clave = m.group(1).upper().replace("_", "")
                desc[clave] = m.group(2).strip()
        return desc

    def _ejecutar_solver(self, tipo, c, restricciones, var_names, texto_orig, metadatos):
        """Ejecuta el solver y muestra resultados."""
        # Leer descripciones del usuario y fusionar con las de la IA (usuario tiene prioridad)
        user_desc = self._leer_descripciones_vars()
        if metadatos is None:
            metadatos = {"descripcion_variables": {}, "fo_descripcion": "",
                         "advertencias": [], "restricciones_desc": []}
        if user_desc:
            metadatos["descripcion_variables"].update(user_desc)

        solver = SimplexSolver(tipo, c, restricciones, nombres_var=var_names)
        record = solver.resolver()
        self.last_record    = record
        self.last_enunciado = texto_orig
        self.last_metadatos = metadatos

        preface = ""
        if metadatos.get("modelo_ia"):
            preface = self._formatear_metadatos_ia(metadatos, tipo, c, restricciones, var_names)

        self._render_output(record, metadatos, preface)
        if OPENPYXL_OK:
            self.btn_exportar.config(state='normal')
        self.btn_guardar_txt.config(state='normal')

        estado_label = {"optima": "OPTIMA", "infactible": "INFACTIBLE",
                        "no_acotada": "NO ACOTADA", "limite": "LIMITE ITERACIONES"}
        n_it = len(record["iteraciones"])
        self.status.config(
            text=f"Listo. Estado: {estado_label.get(record['estado'], '?')}. "
                 f"Iteraciones: {n_it - 1}.",
            fg='#080')

    def _formatear_metadatos_ia(self, metadatos, tipo, c, restricciones, var_names):
        """Genera un preface explicando cómo la IA interpretó el enunciado."""
        out = []
        out.append("=" * 78)
        out.append("  INTERPRETACION DEL ENUNCIADO POR LA IA")
        out.append("=" * 78)
        if metadatos.get("descripcion_variables"):
            out.append("\nVARIABLES DE DECISION:")
            for v in var_names:
                desc = metadatos["descripcion_variables"].get(v, "")
                out.append(f"  {v} = {desc}" if desc else f"  {v}")
        if metadatos.get("fo_descripcion"):
            out.append(f"\nFUNCION OBJETIVO ({metadatos['fo_descripcion']}):")
        else:
            out.append("\nFUNCION OBJETIVO:")
        verbo = "Max" if tipo == "max" else "Min"
        fo_str = " + ".join(f"{fmt_num(ci)} {v}" for ci, v in zip(c, var_names) if ci != 0)
        out.append(f"  {verbo}.  Z = {fo_str}")

        out.append("\nRESTRICCIONES IDENTIFICADAS:")
        descs = metadatos.get("restricciones_desc", [])
        for i, (a, s, b) in enumerate(restricciones):
            izq = " + ".join(f"{fmt_num(ai)} {v}" for ai, v in zip(a, var_names) if ai != 0)
            desc = f"  // {descs[i]}" if i < len(descs) and descs[i] else ""
            out.append(f"  {i+1}) {izq} {s} {fmt_num(b)}{desc}")

        if metadatos.get("advertencias"):
            out.append("\nADVERTENCIAS DE LA IA:")
            for adv in metadatos["advertencias"]:
                out.append(f"  - {adv}")
        out.append("\n  >>> Revisa que la interpretacion sea correcta. Si no, edita el modelo manualmente.")
        return "\n".join(out)

    def on_interpretar_ia(self):
        """Fuerza el uso de la IA (sin auto-detección)."""
        texto = self.txt_input.get('1.0', 'end').strip()
        if not texto:
            messagebox.showwarning("Sin texto", "Por favor pega el enunciado.")
            return
        if not self.api_key:
            messagebox.showinfo(
                "API key requerida",
                "Para usar la IA necesitas una API key de Groq (es GRATIS).\n\n"
                "Obtén una en: https://console.groq.com/keys\n"
                "Luego pulsa 'Configurar API key (IA)'.")
            self.on_config_api()
            if not self.api_key:
                return
        self._resolver_con_ia(texto)

    def on_config_api(self):
        """Diálogo para configurar la API key y el modelo de Groq."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Configurar Groq (API key + modelo)")
        dlg.geometry("620x420")
        dlg.transient(self.root)
        dlg.grab_set()

        tk.Label(dlg, text="API key de Groq",
                 font=('Segoe UI', 11, 'bold')).pack(pady=(12, 4))
        tk.Label(dlg, font=('Segoe UI', 9), justify='left',
                 text=("La API key se guarda localmente en tu computador (no se envia a nadie\n"
                       "excepto a la API oficial de Groq al consultar).\n\n"
                       "Obten una key GRATIS en:  https://console.groq.com/keys   (empieza con 'gsk_').")
                 ).pack(padx=12, pady=(0, 6))

        entry = ttk.Entry(dlg, font=('Consolas', 10), width=72, show='*')
        entry.pack(padx=12, pady=4)
        if self.api_key:
            entry.insert(0, self.api_key)

        show_var = tk.BooleanVar(value=False)
        def toggle_show():
            entry.config(show='' if show_var.get() else '*')
        ttk.Checkbutton(dlg, text="Mostrar key", variable=show_var,
                        command=toggle_show).pack(pady=2)

        # ----- Selector de modelo -----
        tk.Label(dlg, text="Modelo a usar",
                 font=('Segoe UI', 11, 'bold')).pack(pady=(12, 2))
        tk.Label(dlg, font=('Segoe UI', 9), justify='left',
                 text=("Si un modelo da error 403 (Forbidden), prueba otro de la lista o escribe uno manual.\n"
                       "Recomendados (gratis en Groq, buenos razonando):")
                 ).pack(padx=12, pady=(0, 4))

        modelos_sugeridos = [
            "llama-3.3-70b-versatile",
            "llama-3.1-8b-instant",
            "meta-llama/llama-4-scout-17b-16e-instruct",
            "meta-llama/llama-4-maverick-17b-128e-instruct",
            "openai/gpt-oss-120b",
            "openai/gpt-oss-20b",
            "deepseek-r1-distill-llama-70b",
            "qwen/qwen3-32b",
        ]
        model_var = tk.StringVar(value=self.api_model)
        combo = ttk.Combobox(dlg, textvariable=model_var, values=modelos_sugeridos,
                              font=('Consolas', 10), width=60)
        combo.pack(padx=12, pady=2)

        frame_btn = tk.Frame(dlg)
        frame_btn.pack(pady=14)
        def guardar():
            k = entry.get().strip()
            m = model_var.get().strip()
            if not k:
                messagebox.showwarning("Vacio", "Escribe una API key.")
                return
            if not m:
                messagebox.showwarning("Vacio", "Escribe el nombre del modelo.")
                return
            if not k.startswith("gsk_"):
                if not messagebox.askyesno("Formato inusual",
                                            "La key no empieza con 'gsk_'. ¿Guardar de todas formas?"):
                    return
            self.api_key = k
            self.api_model = m
            self.config["groq_api_key"] = k
            self.config["groq_model"] = m
            guardar_config(self.config)
            self.status.config(text=f"Groq configurado. Modelo: {m}", fg='#080')
            dlg.destroy()

        def borrar():
            if messagebox.askyesno("Borrar key",
                                    "¿Borrar la API key guardada en tu computador?"):
                self.api_key = ""
                self.config.pop("groq_api_key", None)
                guardar_config(self.config)
                entry.delete(0, 'end')
                self.status.config(text="API key borrada.")

        ttk.Button(frame_btn, text="Guardar", command=guardar).pack(side='left', padx=4)
        ttk.Button(frame_btn, text="Borrar key guardada", command=borrar).pack(side='left', padx=4)
        ttk.Button(frame_btn, text="Cancelar", command=dlg.destroy).pack(side='left', padx=4)

    def on_ejemplo(self):
        # Menú emergente con los ejemplos
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Taller 1 - Congeladores (Max, con >=)",
                          command=lambda: self._set_input(EJEMPLO_DEFAULT))
        menu.add_command(label="Combustibles (Min, con >= ; Gran M)",
                          command=lambda: self._set_input(EJEMPLO_MIN))
        menu.add_command(label="Pinturas (Max, todas <=)",
                          command=lambda: self._set_input(EJEMPLO_PINTURA))
        try:
            x = self.root.winfo_pointerx()
            y = self.root.winfo_pointery()
            menu.tk_popup(x, y)
        finally:
            menu.grab_release()

    def _set_input(self, texto):
        self.txt_input.delete('1.0', 'end')
        self.txt_input.insert('1.0', texto)

    def on_limpiar(self):
        self.txt_input.delete('1.0', 'end')
        self._set_output("")
        self.last_record = None
        self.btn_exportar.config(state='disabled')
        self.btn_guardar_txt.config(state='disabled')
        self.status.config(text="Listo.", fg='#000')

    def on_cargar_archivo(self):
        fname = filedialog.askopenfilename(
            title="Cargar modelo desde archivo",
            filetypes=[("Archivos de texto", "*.txt"), ("Todos", "*.*")])
        if not fname: return
        try:
            with open(fname, 'r', encoding='utf-8') as f:
                contenido = f.read()
            self._set_input(contenido)
            self.status.config(text=f"Cargado: {os.path.basename(fname)}")
        except Exception as e:
            messagebox.showerror("Error al cargar", str(e))

    def on_exportar(self):
        if not self.last_record:
            messagebox.showwarning("Sin datos", "Primero resuelve un problema.")
            return
        fname = filedialog.asksaveasfilename(
            title="Guardar como Excel",
            defaultextension=".xlsx",
            initialfile=f"simplex_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not fname: return
        try:
            exporter = ExcelExporter()
            exporter.export(self.last_record, enunciado=self.last_enunciado,
                            metadatos=self.last_metadatos, filename=fname)
            self.status.config(text=f"Excel guardado: {fname}", fg='#080')
            messagebox.showinfo("Excel generado",
                                 f"Archivo creado correctamente:\n{fname}")
        except Exception as e:
            messagebox.showerror("Error al exportar", str(e))

    def on_guardar_txt(self):
        if not self.last_record:
            return
        fname = filedialog.asksaveasfilename(
            title="Guardar resultado como texto",
            defaultextension=".txt",
            initialfile=f"simplex_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            filetypes=[("Texto", "*.txt")])
        if not fname: return
        try:
            with open(fname, 'w', encoding='utf-8') as f:
                f.write(format_record_text(self.last_record))
            self.status.config(text=f"Texto guardado: {fname}", fg='#080')
        except Exception as e:
            messagebox.showerror("Error al guardar", str(e))


def main():
    root = tk.Tk()
    app = SimplexApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()